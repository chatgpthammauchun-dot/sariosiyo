"""
╔══════════════════════════════════════════════════════════════╗
║   SARIOSIYO WAREHOUSE BOT v7.0 — To'liq Integratsiya        ║
╠══════════════════════════════════════════════════════════════╣
║  pip install aiogram==3.7.0 sqlalchemy==2.0.30              ║
║              aiosqlite==0.20.0 Pillow==10.3.0               ║
║              qrcode==7.4.2 aiohttp==3.9.5                   ║
║              openpyxl==3.1.2 reportlab==4.2.0               ║
║              python-dotenv==1.0.1                           ║
╠══════════════════════════════════════════════════════════════╣
║  .env:                                                       ║
║    BOT_TOKEN=...                                             ║
║    ADMIN_IDS=123456789,987654321                             ║
║    CHANNEL_ID=@sariosiyo_online                             ║
║    WEBAPP_PUBLIC_URL=https://skaner-2.onrender.com          ║
║    WEBAPP_PORT=8080                                          ║
║    ADMIN_PASSWORDS=admin123,sariosiyo                       ║
╚══════════════════════════════════════════════════════════════╝
"""

import asyncio, calendar, hashlib, io, json, logging, os, random
import shutil, string, time
from collections import defaultdict
from datetime import datetime, date
from typing import Optional

from dotenv import load_dotenv
load_dotenv()

# ── Config ──────────────────────────────────────────────────────────────────
BOT_TOKEN         = os.getenv("BOT_TOKEN", "")
ADMIN_IDS         = [int(x) for x in os.getenv("ADMIN_IDS","").split(",") if x.strip().isdigit()]
CHANNEL_ID        = os.getenv("CHANNEL_ID","@sariosiyo_online")
WEBAPP_PUBLIC_URL = os.getenv("WEBAPP_PUBLIC_URL", f"http://localhost:{os.getenv('WEBAPP_PORT','8080')}")
WEBAPP_PORT       = int(os.getenv("WEBAPP_PORT","8080"))
ADMIN_PASSWORDS   = [p.strip() for p in os.getenv("ADMIN_PASSWORDS","admin123,sariosiyo").split(",") if p.strip()]
LOW_STOCK_ALERT   = int(os.getenv("LOW_STOCK_ALERT","5"))
DAILY_REPORT_HOUR = int(os.getenv("DAILY_REPORT_HOUR","21"))
DB_PATH           = "warehouse.db"
BACKUP_DIR        = "backups"
MEDIA_DIR         = "media"

os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(MEDIA_DIR, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

# ── SQLAlchemy ───────────────────────────────────────────────────────────────
from sqlalchemy import (
    Boolean, Column, DateTime, Float, ForeignKey,
    Integer, String, Text, func, select, extract
)
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.orm import DeclarativeBase, relationship

engine = create_async_engine(f"sqlite+aiosqlite:///{DB_PATH}", echo=False)
AsyncSessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

try:
    from PIL import Image, ImageDraw, ImageFont
    PILLOW_OK = True
except ImportError:
    PILLOW_OK = False

try:
    import qrcode as qrlib
    QR_OK = True
except ImportError:
    QR_OK = False

# ════════════════════════════════════════════════════════════════
#  1. MODELS
# ════════════════════════════════════════════════════════════════
class Base(DeclarativeBase): pass

class Worker(Base):
    __tablename__ = "workers"
    id          = Column(Integer, primary_key=True, autoincrement=True)
    telegram_id = Column(Integer, unique=True, index=True, nullable=False)
    name        = Column(String(128), nullable=False)
    phone       = Column(String(20), default="")
    role        = Column(String(20), default="seller")   # admin | seller
    is_active   = Column(Boolean, default=True)
    created_at  = Column(DateTime, default=datetime.utcnow)
    sales       = relationship("Sale", back_populates="worker")

class Product(Base):
    __tablename__ = "products"
    id            = Column(Integer, primary_key=True, autoincrement=True)
    code          = Column(String(50), unique=True, index=True, nullable=False)
    barcode       = Column(String(50), default="", index=True)
    name          = Column(String(256), nullable=False)
    category      = Column(String(128), default="")
    description   = Column(Text, default="")
    color         = Column(String(64), default="")
    size          = Column(String(32), default="")
    unit          = Column(String(20), default="dona")
    buy_price     = Column(Float, default=0.0)
    sell_price    = Column(Float, default=0.0)
    stock         = Column(Integer, default=0)
    min_stock     = Column(Integer, default=5)
    photo_file_id = Column(String(256), default="")
    photo_url     = Column(String(512), default="")   # public CDN url (ixtiyoriy)
    is_active     = Column(Boolean, default=True)
    created_at    = Column(DateTime, default=datetime.utcnow)
    updated_at    = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    sales         = relationship("Sale",   back_populates="product")
    supplies      = relationship("Supply", back_populates="product")

class Sale(Base):
    __tablename__ = "sales"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    product_id = Column(Integer, ForeignKey("products.id"), nullable=False)
    worker_id  = Column(Integer, ForeignKey("workers.id"), nullable=True)
    quantity   = Column(Integer, default=1)
    sell_price = Column(Float, default=0.0)
    buy_price  = Column(Float, default=0.0)
    profit     = Column(Float, default=0.0)
    note       = Column(String(256), default="")
    sold_at    = Column(DateTime, default=datetime.utcnow)
    product    = relationship("Product", back_populates="sales")
    worker     = relationship("Worker",  back_populates="sales")

class Supply(Base):
    __tablename__ = "supplies"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    product_id = Column(Integer, ForeignKey("products.id"), nullable=False)
    worker_id  = Column(Integer, ForeignKey("workers.id"), nullable=True)
    quantity   = Column(Integer, default=0)
    buy_price  = Column(Float, default=0.0)
    note       = Column(String(256), default="")
    added_at   = Column(DateTime, default=datetime.utcnow)
    product    = relationship("Product", back_populates="supplies")
    worker     = relationship("Worker")

class ChannelPost(Base):
    __tablename__ = "channel_posts"
    id         = Column(Integer, primary_key=True, autoincrement=True)
    product_id = Column(Integer, ForeignKey("products.id"))
    message_id = Column(Integer, default=0)
    caption    = Column(Text, default="")
    posted_by  = Column(Integer, default=0)
    posted_at  = Column(DateTime, default=datetime.utcnow)

# ════════════════════════════════════════════════════════════════
#  2. HELPERS
# ════════════════════════════════════════════════════════════════
def _gen_barcode() -> str:
    return "SAR-" + "".join(random.choices(string.digits, k=6))

async def unique_barcode(session: AsyncSession) -> str:
    for _ in range(200):
        code = _gen_barcode()
        r = await session.execute(select(Product).where(Product.barcode == code))
        if not r.scalar_one_or_none():
            return code
    return "SAR-" + "".join(random.choices(string.digits, k=8))

async def get_product(session: AsyncSession, code: str) -> Optional[Product]:
    code = code.strip().upper()
    r = await session.execute(
        select(Product).where(
            Product.is_active == True,
            (func.upper(Product.code)    == code) |
            (func.upper(Product.barcode) == code)
        )
    )
    return r.scalar_one_or_none()

async def search_products(session: AsyncSession, q: str) -> list:
    pat = f"%{q.strip().upper()}%"
    r = await session.execute(
        select(Product).where(
            Product.is_active == True,
            (func.upper(Product.code).like(pat))    |
            (func.upper(Product.name).like(pat))    |
            (func.upper(Product.barcode).like(pat)) |
            (func.upper(Product.category).like(pat))
        ).order_by(Product.name).limit(20)
    )
    return r.scalars().all()

async def get_worker(session: AsyncSession, tid: int) -> Optional[Worker]:
    r = await session.execute(
        select(Worker).where(Worker.telegram_id == tid, Worker.is_active == True)
    )
    return r.scalar_one_or_none()

async def is_admin(session: AsyncSession, tid: int) -> bool:
    if tid in ADMIN_IDS: return True
    w = await get_worker(session, tid)
    return w is not None and w.role == "admin"

async def get_low_stock(session: AsyncSession) -> list:
    r = await session.execute(
        select(Product).where(Product.is_active == True, Product.stock <= Product.min_stock)
        .order_by(Product.stock)
    )
    return r.scalars().all()

async def do_sale(session, product, qty, worker, price=None, note=""):
    sp     = price or product.sell_price
    profit = (sp - product.buy_price) * qty
    sale   = Sale(
        product_id=product.id, worker_id=worker.id if worker else None,
        quantity=qty, sell_price=sp, buy_price=product.buy_price,
        profit=profit, note=note
    )
    product.stock -= qty
    product.updated_at = datetime.utcnow()
    session.add_all([sale, product])
    await session.commit()
    return sale

async def do_supply(session, product, qty, buy_price, worker, note=""):
    sup = Supply(
        product_id=product.id, worker_id=worker.id if worker else None,
        quantity=qty, buy_price=buy_price, note=note
    )
    product.stock    += qty
    product.buy_price = buy_price
    product.updated_at = datetime.utcnow()
    session.add_all([sup, product])
    await session.commit()
    return sup

def fmt(n: float) -> str:
    return f"{round(n):,}".replace(",", " ")

def product_text(p: Product, admin: bool = True) -> str:
    icon  = "🔴" if p.stock == 0 else ("🟡" if p.stock <= p.min_stock else "🟢")
    lines = [f"📦 <b>{p.name}</b>"]
    if p.barcode:  lines.append(f"📊 Barcode: <code>{p.barcode}</code>")
    lines.append(f"🔖 Kod: <code>{p.code}</code>")
    if p.category: lines.append(f"🗂 {p.category}")
    if p.color:    lines.append(f"🎨 {p.color}")
    if p.size:     lines.append(f"📐 {p.size}")
    lines.append(f"📏 {p.unit}")
    if admin:
        profit = p.sell_price - p.buy_price
        pct    = (profit / p.buy_price * 100) if p.buy_price > 0 else 0
        lines.append(f"💰 Tannarx: {fmt(p.buy_price)} so'm")
        lines.append(f"💵 Sotish: <b>{fmt(p.sell_price)} so'm</b>")
        lines.append(f"📈 Foyda: {fmt(profit)} so'm ({pct:.0f}%)")
    else:
        lines.append(f"💵 Narx: <b>{fmt(p.sell_price)} so'm</b>")
    lines.append(f"{icon} Stok: <b>{p.stock} {p.unit}</b>")
    if p.description: lines.append(f"\n📄 {p.description}")
    return "\n".join(lines)

def channel_caption(p: Product) -> str:
    lines = [f"🛍 <b>{p.name}</b>", ""]
    if p.color: lines.append(f"🎨 Rang: {p.color}")
    if p.size:  lines.append(f"📐 O'lcham: {p.size}")
    lines.append(f"📦 Birlik: {p.unit}")
    lines += ["", f"💵 Narx: <b>{fmt(p.sell_price)} so'm</b>", ""]
    lines += [f"🔖 Kod: <code>{p.code}</code>", ""]
    lines += ["📞 +998909182186", "📞 +998947043111", "", f"🌐 {CHANNEL_ID}"]
    return "\n".join(lines)

def public_url(p: Product) -> str:
    return f"{WEBAPP_PUBLIC_URL}/p/{p.barcode or p.code}"

def _progress(step: int, total: int = 11) -> str:
    return f"[{'█'*step}{'░'*(total-step)}] {step}/{total}"

# ════════════════════════════════════════════════════════════════
#  3. QR / LABEL IMAGE GENERATION
# ════════════════════════════════════════════════════════════════
_RED   = (220, 30, 50)
_DARK  = (25, 25, 25)
_WHITE = (255, 255, 255)
_GREEN = (80, 200, 100)
_GRAY  = (180, 180, 180)
_GOLD  = (212, 175, 55)

def _font(size: int, bold: bool = True):
    paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold
            else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
    ]
    for p in paths:
        if os.path.exists(p):
            try: return ImageFont.truetype(p, size)
            except: pass
    return ImageFont.load_default()

def _centered_text(draw, y, txt, font, color, width):
    try:
        bb  = draw.textbbox((0,0), txt, font=font)
        x   = (width - (bb[2]-bb[0])) // 2
        draw.text((x, y), txt, fill=color, font=font)
    except:
        pass

def _make_qr_image(data: str, box_size=10, border=4, version=3,
                   error=None, fill="#1a1a2e"):
    if not QR_OK: return None
    try:
        import qrcode
        ec = error or qrcode.constants.ERROR_CORRECT_H
        qr = qrcode.QRCode(version=version, error_correction=ec,
                           box_size=box_size, border=border)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color=fill, back_color="white")
        return img.get_image()
    except:
        return None

def make_product_qr(p: Product) -> bytes:
    """Chiroyli QR label — bot public sahifa URL bilan"""
    if not (QR_OK and PILLOW_OK): return b""
    try:
        import qrcode
        url      = public_url(p)
        barcode  = p.barcode or p.code
        qr_img   = _make_qr_image(url, box_size=10, border=4, version=3,
                                   error=qrcode.constants.ERROR_CORRECT_H)
        if not qr_img: return b""

        QW, QH = qr_img.size
        HDR    = 44
        FTR    = 28 + (22 if p.name else 0) + (24 if p.sell_price>0 else 0) + \
                 (20 if (p.color or p.size) else 0) + (18 if p.category else 0) + 22
        W = QW + 24
        H = HDR + QH + 10 + FTR

        canvas = Image.new("RGB", (W, H), _WHITE)
        draw   = ImageDraw.Draw(canvas)
        draw.rectangle([0, 0, W, HDR], fill=_RED)
        _centered_text(draw, 10, "SARIOSIYO ONLINE", _font(20, True), _WHITE, W)

        qt = HDR + 5
        canvas.paste(qr_img, (12, qt))
        draw.rectangle([10, qt-2, 10+QW+3, qt+QH+2], outline=_GRAY, width=1)

        fy = qt + QH + 8
        draw.rectangle([0, fy-2, W, fy], fill=_RED)
        _centered_text(draw, fy+4, barcode, _font(18, True), _DARK, W); fy += 28
        if p.name:
            nm = p.name[:30] if len(p.name)>30 else p.name
            _centered_text(draw, fy, nm, _font(16, False), (50,50,80), W); fy += 22
        if p.sell_price > 0:
            _centered_text(draw, fy, f"{fmt(p.sell_price)} so'm", _font(18, True), _RED, W); fy += 24
        extras = [x for x in [p.color, p.size] if x]
        if extras:
            _centered_text(draw, fy, " · ".join(extras), _font(14, False), _GRAY, W); fy += 20
        if p.category:
            _centered_text(draw, fy, p.category, _font(13, False), (170,170,170), W)

        buf = io.BytesIO()
        canvas.save(buf, "PNG", optimize=True)
        return buf.getvalue()
    except Exception as e:
        logger.error(f"QR xato: {e}")
        return b""

def make_label_58mm(p: Product) -> bytes:
    """58mm termal printer label"""
    if not (PILLOW_OK and QR_OK): return b""
    try:
        import qrcode
        W, H   = 576, 320
        img    = Image.new("RGB", (W, H), _WHITE)
        draw   = ImageDraw.Draw(img)
        draw.rectangle([0, 0, W, 36], fill=_RED)
        _centered_text(draw, 7, "SARIOSIYO ONLINE", _font(17, True), _WHITE, W)

        qr_img = _make_qr_image(public_url(p), box_size=5, border=2, version=1,
                                 error=qrcode.constants.ERROR_CORRECT_M)
        QSZ = 210
        if qr_img:
            qr_img = qr_img.resize((QSZ, QSZ), Image.LANCZOS)
            img.paste(qr_img, (8, 42))

        rx, ry = QSZ+20, 46
        draw.text((rx, ry), p.barcode or p.code, fill=_DARK, font=_font(15, True)); ry += 24
        nm = p.name[:18]
        draw.text((rx, ry), nm, fill=(50,50,80), font=_font(14, False)); ry += 22
        draw.text((rx, ry), f"{fmt(p.sell_price)} so'm", fill=_RED, font=_font(20, True)); ry += 28
        extras = [x for x in [p.color, p.size] if x]
        if extras: draw.text((rx, ry), " | ".join(extras), fill=_GRAY, font=_font(13, False))

        cy = 268
        for x in range(0, W, 10):
            draw.line([(x, cy), (x+5, cy)], fill=_GRAY, width=1)
        draw.text((6, cy+5), p.barcode or p.code, fill=(170,170,170), font=_font(13, False))
        draw.rectangle([0, H-6, W, H], fill=_RED)

        buf = io.BytesIO()
        img.save(buf, "PNG", dpi=(203,203))
        return buf.getvalue()
    except Exception as e:
        logger.error(f"Label 58mm xato: {e}")
        return b""

def make_qr_sheet(items: list, cols: int = 3) -> bytes:
    """A4 QR varaq"""
    if not (PILLOW_OK and QR_OK): return b""
    try:
        IW, IH = 280, 330
        PAD    = 20
        rows   = (len(items)+cols-1)//cols
        SW     = cols*IW + (cols+1)*PAD
        SH     = rows*IH + (rows+1)*PAD + 60
        sheet  = Image.new("RGB", (SW, SH), "white")
        draw   = ImageDraw.Draw(sheet)
        draw.rectangle([0, 0, SW, 55], fill=_RED)
        _centered_text(draw, 14, "SARIOSIYO ONLINE — QR KODLAR", _font(22, True), _WHITE, SW)
        for idx, (bc, name) in enumerate(items):
            row = idx//cols; col = idx%cols
            x   = PAD + col*(IW+PAD)
            y   = 60 + PAD + row*(IH+PAD)
            # create a temporary product-like object for make_product_qr
            class _P:
                barcode=bc; code=bc; name_=name; sell_price=0; color=""; size=""; category=""
            qb = _make_qr_image(f"{WEBAPP_PUBLIC_URL}/p/{bc}", box_size=7, border=3, version=2)
            if qb:
                qr_img = qb.resize((IW-10, IH-60), Image.LANCZOS)
                sheet.paste(qr_img, (x+5, y+5))
            # barcode text
            draw.text((x+5, y+IH-50), bc, fill=_DARK, font=_font(16, True))
            nm = name[:20] if len(name)>20 else name
            draw.text((x+5, y+IH-28), nm, fill=(80,80,80), font=_font(13, False))
            draw.rectangle([x, y, x+IW, y+IH], outline=(220,220,220), width=1)
        buf = io.BytesIO()
        sheet.save(buf, "PNG", dpi=(300,300))
        return buf.getvalue()
    except Exception as e:
        logger.error(f"QR sheet xato: {e}")
        return b""

# ════════════════════════════════════════════════════════════════
#  4. HTML PAGES
# ════════════════════════════════════════════════════════════════
def get_public_page_html(p: Product) -> str:
    color_b = f"<div class='info'><span class='label'>🎨 Rang</span><span class='val'>{p.color}</span></div>" if p.color else ""
    size_b  = f"<div class='info'><span class='label'>📐 O'lcham</span><span class='val'>{p.size}</span></div>" if p.size else ""
    cat_b   = f"<div class='info'><span class='label'>🗂 Kategoriya</span><span class='val'>{p.category}</span></div>" if p.category else ""
    unit_b  = f"<div class='info'><span class='label'>📦 Birlik</span><span class='val'>{p.unit}</span></div>"
    desc_b  = f"<div class='desc'>{p.description}</div>" if p.description else ""
    img_b   = f'<img class="pimg" src="{p.photo_url}" onerror="this.outerHTML=\'<div class=nimg>📦</div>\'">' if p.photo_url else "<div class='nimg'>📦</div>"
    channel = CHANNEL_ID.lstrip("@")
    barcode = p.barcode or p.code

    stock_cls = "s-empty" if p.stock==0 else ("s-low" if p.stock<=p.min_stock else "s-ok")
    stock_txt = "🔴 Tugagan" if p.stock==0 else (f"🟡 Kam qoldi" if p.stock<=p.min_stock else "🟢 Mavjud")

    return f"""<!DOCTYPE html>
<html lang="uz">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,user-scalable=no">
<meta property="og:title" content="{p.name} — Sariosiyo Online">
<meta property="og:description" content="Narx: {fmt(p.sell_price)} so'm">
<title>{p.name} — Sariosiyo Online</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700;800&display=swap');
:root{{--red:#E8192C;--red2:#C0141F;--dark:#0D0D14;--card:#13131E;--card2:#1A1A2E;
  --border:rgba(255,255,255,0.07);--white:#F4F6FB;--gray:#8892A4;
  --green:#22C55E;--yellow:#F59E0B;--blue:#3B82F6}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Space Grotesk',sans-serif;background:var(--dark);color:var(--white);min-height:100vh}}
.header{{background:linear-gradient(135deg,var(--red2),var(--red));padding:0 20px;height:60px;
  display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10;
  box-shadow:0 4px 30px rgba(232,25,44,.4)}}
.h-left{{display:flex;align-items:center;gap:12px}}
.logo{{width:38px;height:38px;background:rgba(255,255,255,.15);border-radius:11px;
  display:flex;align-items:center;justify-content:center;font-size:19px;border:1px solid rgba(255,255,255,.2)}}
.h-title{{font-size:.95rem;font-weight:700}}
.h-sub{{font-size:.7rem;opacity:.85}}
.admin-btn{{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:white;
  padding:7px 14px;border-radius:20px;font-family:inherit;font-size:.8rem;font-weight:600;cursor:pointer}}
.wrap{{max-width:480px;margin:0 auto;padding:20px 16px 50px}}
.card{{background:var(--card);border-radius:24px;overflow:hidden;border:1px solid var(--border);
  box-shadow:0 20px 60px rgba(0,0,0,.5);margin-bottom:16px}}
.pimg{{width:100%;height:280px;object-fit:cover;display:block}}
.nimg{{width:100%;height:220px;background:linear-gradient(135deg,#0A0A14,#13131E);
  display:flex;align-items:center;justify-content:center;font-size:5rem}}
.body{{padding:22px}}
.badge{{display:inline-block;background:var(--red);color:white;font-size:.7rem;font-weight:700;
  padding:3px 10px;border-radius:20px;margin-bottom:10px}}
.name{{font-size:1.35rem;font-weight:800;line-height:1.3;margin-bottom:5px}}
.code{{font-size:.76rem;color:var(--gray);font-family:monospace;margin-bottom:14px}}
.price{{font-size:2rem;font-weight:800;color:var(--green);margin-bottom:14px;
  display:flex;align-items:baseline;gap:7px}}
.price span{{font-size:.9rem;color:var(--gray);font-weight:400}}
.stock-badge{{display:inline-flex;align-items:center;gap:5px;padding:5px 12px;border-radius:20px;
  font-size:.8rem;font-weight:600;margin-bottom:14px}}
.s-ok{{background:rgba(34,197,94,.15);color:var(--green)}}
.s-low{{background:rgba(245,158,11,.15);color:var(--yellow)}}
.s-empty{{background:rgba(232,25,44,.15);color:var(--red)}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:14px}}
.info{{background:rgba(255,255,255,.04);border-radius:11px;padding:10px 12px}}
.label{{display:block;font-size:.67rem;color:var(--gray);text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px}}
.val{{font-size:.9rem;font-weight:600}}
.desc{{font-size:.86rem;color:var(--gray);line-height:1.65;margin-bottom:14px;
  padding:12px;background:rgba(255,255,255,.03);border-radius:11px}}
.divider{{height:1px;background:var(--border);margin:14px 0}}
.contacts{{font-size:.87rem;color:var(--gray);line-height:2.1}}
.contacts a{{color:var(--white);text-decoration:none}}
.btn{{display:flex;align-items:center;justify-content:center;gap:7px;
  padding:14px;border-radius:13px;font-family:inherit;font-size:.97rem;
  font-weight:700;cursor:pointer;border:none;text-decoration:none;transition:all .2s;
  width:100%;margin-top:10px}}
.btn-red{{background:linear-gradient(135deg,var(--red),var(--red2));color:white;
  box-shadow:0 4px 20px rgba(232,25,44,.3)}}
.btn-tg{{background:linear-gradient(135deg,#2481CC,#1a6ba8);color:white}}
.btn:active{{opacity:.85;transform:scale(.98)}}
/* Admin panel */
.admin-panel{{display:none;background:var(--card);border-radius:24px;border:1px solid rgba(232,25,44,.3);
  overflow:hidden;margin-bottom:16px;animation:fu .35s cubic-bezier(.16,1,.3,1)}}
.admin-panel.show{{display:block}}
.ap-header{{background:linear-gradient(135deg,rgba(232,25,44,.2),rgba(232,25,44,.05));
  border-bottom:1px solid rgba(232,25,44,.2);padding:14px 20px;
  display:flex;align-items:center;justify-content:space-between}}
.ap-title{{font-size:.88rem;font-weight:700;color:var(--red);display:flex;align-items:center;gap:7px}}
.ap-badge{{background:var(--red);color:white;font-size:.63rem;font-weight:700;padding:2px 7px;border-radius:9px}}
.ap-body{{padding:18px 20px}}
.stats{{display:grid;grid-template-columns:1fr 1fr;gap:9px;margin-bottom:14px}}
.stat{{background:rgba(255,255,255,.04);border:1px solid var(--border);border-radius:13px;padding:13px}}
.sl{{font-size:.66rem;color:var(--gray);text-transform:uppercase;letter-spacing:.7px;margin-bottom:5px}}
.sv{{font-size:1.05rem;font-weight:700}}
.sv.g{{color:var(--green)}}.sv.r{{color:var(--red)}}.sv.y{{color:var(--yellow)}}.sv.b{{color:var(--blue)}}
.pbar-wrap{{margin:12px 0}}
.pbar-label{{display:flex;justify-content:space-between;font-size:.76rem;color:var(--gray);margin-bottom:5px}}
.pbar-bg{{height:7px;background:rgba(255,255,255,.07);border-radius:8px;overflow:hidden}}
.pbar-fill{{height:100%;background:linear-gradient(90deg,var(--green),#16a34a);border-radius:8px;transition:width 1s ease}}
.info-box{{background:rgba(255,255,255,.03);border:1px solid var(--border);border-radius:11px;
  padding:11px 13px;font-size:.81rem;color:var(--gray);line-height:1.9;margin-bottom:14px}}
.ap-actions{{display:grid;grid-template-columns:1fr 1fr;gap:7px}}
.ap-btn{{background:rgba(255,255,255,.05);border:1px solid var(--border);color:var(--white);
  font-family:inherit;font-size:.8rem;font-weight:600;padding:9px;border-radius:11px;
  cursor:pointer;transition:all .2s;display:flex;align-items:center;justify-content:center;gap:5px}}
.ap-btn:hover{{background:rgba(255,255,255,.1)}}
.ap-btn.danger{{color:var(--red);border-color:rgba(232,25,44,.3)}}
/* Modals */
.overlay{{position:fixed;inset:0;background:rgba(0,0,0,.87);z-index:200;
  display:flex;align-items:flex-end;justify-content:center;
  opacity:0;pointer-events:none;transition:opacity .3s}}
.overlay.open{{opacity:1;pointer-events:all}}
.modal{{background:var(--card2);border-radius:26px 26px 0 0;width:100%;max-width:480px;
  padding:8px 22px 38px;border-top:1px solid var(--border);
  transform:translateY(100%);transition:transform .38s cubic-bezier(.16,1,.3,1);
  max-height:88vh;overflow-y:auto}}
.overlay.open .modal{{transform:translateY(0)}}
.mhandle{{width:38px;height:4px;background:rgba(255,255,255,.15);border-radius:4px;margin:10px auto 18px}}
.mtitle{{font-size:1.15rem;font-weight:800;text-align:center;margin-bottom:5px}}
.msub{{font-size:.83rem;color:var(--gray);text-align:center;margin-bottom:20px}}
.role-grid{{display:grid;grid-template-columns:1fr 1fr;gap:11px;margin-bottom:18px}}
.role-card{{background:rgba(255,255,255,.04);border:2px solid var(--border);border-radius:17px;
  padding:18px 12px;text-align:center;cursor:pointer;transition:all .22s}}
.role-card.selected{{border-color:var(--red);background:rgba(232,25,44,.1)}}
.ri{{font-size:2.3rem;margin-bottom:8px}}
.rn{{font-size:.92rem;font-weight:700;margin-bottom:3px}}
.rd{{font-size:.7rem;color:var(--gray)}}
.pass-step{{display:none}}.pass-step.show{{display:block}}
.pass-wrap{{position:relative;margin-bottom:12px}}
.pass-input{{width:100%;background:rgba(255,255,255,.06);border:1.5px solid var(--border);
  color:var(--white);font-family:inherit;font-size:1.05rem;font-weight:600;
  padding:13px 46px 13px 17px;border-radius:13px;outline:none;
  letter-spacing:3px;text-align:center;transition:border-color .2s}}
.pass-input:focus{{border-color:var(--red)}}
.pass-input.err{{border-color:var(--red);animation:shake .4s ease}}
.pass-toggle{{position:absolute;right:13px;top:50%;transform:translateY(-50%);
  background:none;border:none;color:var(--gray);cursor:pointer;font-size:1rem}}
.pass-hint{{font-size:.78rem;color:var(--gray);text-align:center;margin-bottom:14px}}
.pass-err{{font-size:.8rem;color:var(--red);text-align:center;margin-bottom:9px;display:none}}
.pass-err.show{{display:block}}
.mbtn{{width:100%;padding:14px;border-radius:13px;font-family:inherit;font-size:.97rem;
  font-weight:700;cursor:pointer;border:none;transition:all .2s;margin-bottom:9px}}
.mbtn-p{{background:linear-gradient(135deg,var(--red),var(--red2));color:white}}
.mbtn-s{{background:rgba(255,255,255,.05);color:var(--gray);border:1px solid var(--border)}}
/* Action modal */
.aoverlay{{position:fixed;inset:0;background:rgba(0,0,0,.87);z-index:300;
  display:flex;align-items:flex-end;justify-content:center;
  opacity:0;pointer-events:none;transition:opacity .3s}}
.aoverlay.open{{opacity:1;pointer-events:all}}
.amodal{{background:var(--card2);border-radius:26px 26px 0 0;width:100%;max-width:480px;
  padding:8px 22px 38px;border-top:1px solid var(--border);
  transform:translateY(100%);transition:transform .38s cubic-bezier(.16,1,.3,1);
  max-height:88vh;overflow-y:auto}}
.aoverlay.open .amodal{{transform:translateY(0)}}
.igroup{{margin-bottom:13px}}
.ilabel{{font-size:.76rem;color:var(--gray);margin-bottom:5px;display:block;
  text-transform:uppercase;letter-spacing:.5px}}
.finput{{width:100%;background:rgba(255,255,255,.06);border:1.5px solid var(--border);
  color:var(--white);font-family:inherit;font-size:.93rem;
  padding:11px 15px;border-radius:11px;outline:none;transition:border-color .2s}}
.finput:focus{{border-color:var(--red)}}
.finput option{{background:#1A1A2E;color:var(--white)}}
.rbox{{background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.2);
  border-radius:11px;padding:13px;margin:13px 0;font-size:.88rem;line-height:1.8}}
.rbox.warn{{background:rgba(245,158,11,.1);border-color:rgba(245,158,11,.2)}}
.rl{{font-size:.7rem;color:var(--gray);text-transform:uppercase;letter-spacing:.5px}}
.rv{{font-size:1.05rem;font-weight:700;color:var(--green)}}
.rv.y{{color:var(--yellow)}}
.hist-item{{background:rgba(255,255,255,.04);border:1px solid var(--border);border-radius:11px;
  padding:11px 13px;margin-bottom:7px;display:flex;justify-content:space-between;align-items:center}}
.hi-l{{font-size:.83rem}}.hi-d{{font-size:.7rem;color:var(--gray);margin-top:2px}}
.hi-a{{font-weight:700}}.hi-a.g{{color:var(--green)}}.hi-a.r{{color:var(--red)}}
.esect{{font-size:.76rem;color:var(--red);font-weight:700;text-transform:uppercase;
  letter-spacing:.7px;margin:14px 0 9px}}
/* Toast */
.toast{{position:fixed;bottom:28px;left:50%;transform:translateX(-50%) translateY(100px);
  background:#22C55E;color:white;padding:11px 22px;border-radius:20px;
  font-size:.88rem;font-weight:700;z-index:999;
  transition:transform .3s cubic-bezier(.16,1,.3,1);white-space:nowrap;
  box-shadow:0 4px 20px rgba(0,0,0,.3)}}
.toast.show{{transform:translateX(-50%) translateY(0)}}
.toast.err{{background:var(--red)}}.toast.inf{{background:var(--blue)}}
/* Confirm */
.cov{{position:fixed;inset:0;background:rgba(0,0,0,.87);z-index:400;
  display:flex;align-items:center;justify-content:center;padding:20px;
  opacity:0;pointer-events:none;transition:opacity .2s}}
.cov.open{{opacity:1;pointer-events:all}}
.cbox{{background:var(--card2);border-radius:22px;padding:26px 22px;
  width:100%;max-width:330px;border:1px solid var(--border);text-align:center;
  transform:scale(.9);transition:transform .22s cubic-bezier(.16,1,.3,1)}}
.cov.open .cbox{{transform:scale(1)}}
.ci{{font-size:2.8rem;margin-bottom:12px}}
.ct{{font-size:1.05rem;font-weight:800;margin-bottom:7px}}
.cm{{font-size:.83rem;color:var(--gray);margin-bottom:20px;line-height:1.55}}
.cbtns{{display:flex;gap:9px}}
.cbtn{{flex:1;padding:12px;border-radius:11px;font-family:inherit;font-size:.92rem;font-weight:700;cursor:pointer;border:none}}
.cbtn-c{{background:rgba(255,255,255,.07);color:var(--gray);border:1px solid var(--border)}}
.cbtn-ok{{background:linear-gradient(135deg,var(--red),var(--red2));color:white}}
/* QR section */
.qr-sec{{background:var(--card);border:1px solid var(--border);border-radius:20px;
  padding:18px;text-align:center;margin-bottom:16px}}
.qr-sec h3{{font-size:.88rem;font-weight:700;color:var(--gray);margin-bottom:12px}}
.qr-img{{width:156px;height:156px;background:white;border-radius:13px;
  margin:0 auto 11px;display:flex;align-items:center;justify-content:center;overflow:hidden}}
.qr-img img{{width:100%;height:100%}}
.qr-code{{font-family:monospace;font-size:.83rem;color:var(--gray);
  background:rgba(255,255,255,.04);padding:5px 13px;border-radius:7px;display:inline-block}}
footer{{text-align:center;padding:14px;color:rgba(255,255,255,.18);font-size:.74rem}}
@keyframes fu{{from{{opacity:0;transform:translateY(18px)}}to{{opacity:1;transform:translateY(0)}}}}
@keyframes shake{{0%,100%{{transform:translateX(0)}}20%,60%{{transform:translateX(-5px)}}40%,80%{{transform:translateX(5px)}}}}
</style>
</head>
<body>

<div class="header">
  <div class="h-left">
    <div class="logo">🏪</div>
    <div><div class="h-title">Sariosiyo Online</div><div class="h-sub">Rasmiy do'kon</div></div>
  </div>
  <button class="admin-btn" onclick="openModal()">🔐 Admin</button>
</div>

<div class="wrap">

  <!-- ADMIN PANEL -->
  <div class="admin-panel" id="adminPanel">
    <div class="ap-header">
      <div class="ap-title">🔐 Admin Panel <span class="ap-badge">MAXFIY</span></div>
      <button onclick="logout()" style="background:none;border:none;color:var(--gray);cursor:pointer;font-size:.8rem">Chiqish ✕</button>
    </div>
    <div class="ap-body">
      <div class="stats">
        <div class="stat"><div class="sl">Tannarx</div><div class="sv r" id="sBuy">{fmt(p.buy_price)} so'm</div></div>
        <div class="stat"><div class="sl">Sotish narxi</div><div class="sv g" id="sSell">{fmt(p.sell_price)} so'm</div></div>
        <div class="stat"><div class="sl">Sof foyda</div><div class="sv y" id="sProfit">+{fmt(p.sell_price-p.buy_price)} so'm</div></div>
        <div class="stat"><div class="sl">Stok</div><div class="sv b" id="sStock">{p.stock} {p.unit}</div></div>
      </div>
      <div class="pbar-wrap">
        <div class="pbar-label"><span>Foyda foizi</span><span style="color:var(--green);font-weight:700" id="sPct">{round((p.sell_price-p.buy_price)/p.buy_price*100) if p.buy_price>0 else 0}%</span></div>
        <div class="pbar-bg"><div class="pbar-fill" id="pbar" style="width:{min(round((p.sell_price-p.buy_price)/p.buy_price*100) if p.buy_price>0 else 0, 100)}%"></div></div>
      </div>
      <div class="info-box">
        📦 Omborda: <b style="color:var(--white)">{p.stock} {p.unit}</b><br>
        📅 Barcode: <b style="color:var(--white)">{p.barcode or p.code}</b><br>
        🔗 Public link: <a href="{public_url(p)}" style="color:var(--blue);font-size:.8rem">{public_url(p)}</a>
      </div>
      <div class="ap-actions">
        <button class="ap-btn" onclick="openSotish()">📤 Sotish</button>
        <button class="ap-btn" onclick="openKirim()">📥 Kirim</button>
        <button class="ap-btn" onclick="openTahrir()">✏️ Tahrirlash</button>
        <button class="ap-btn" onclick="openTarix()">📊 Tarix</button>
        <button class="ap-btn danger" onclick="confirmArchive()">🗑 Arxivlash</button>
        <button class="ap-btn" onclick="copyLink()">🔗 Link nusxa</button>
      </div>
    </div>
  </div>

  <!-- PRODUCT CARD -->
  <div class="card">
    <div style="position:relative">
      {img_b}
      <div style="position:absolute;top:13px;left:13px;background:var(--red);color:white;
        font-size:.7rem;font-weight:700;padding:3px 10px;border-radius:18px">{p.category or 'Mahsulot'}</div>
      <div style="position:absolute;bottom:13px;right:13px;background:rgba(0,0,0,.7);
        color:rgba(255,255,255,.75);font-size:.7rem;font-family:monospace;padding:3px 9px;border-radius:7px">{barcode}</div>
    </div>
    <div class="body">
      <div class="badge">{p.category or 'Mahsulot'}</div>
      <div class="name">{p.name}</div>
      <div class="code">{p.code} · {barcode}</div>
      <div class="price">{fmt(p.sell_price)} <span>so'm</span></div>
      <div class="stock-badge {stock_cls}" id="stockBadge">{stock_txt}</div>
      <div class="grid">{color_b}{size_b}{cat_b}{unit_b}</div>
      {desc_b}
      <div class="divider"></div>
      <div class="contacts">
        📱 <a href="tel:+998909182186">+998 90 918 21 86</a><br>
        📱 <a href="tel:+998947043111">+998 94 704 31 11</a><br>
        🌐 <a href="https://t.me/{channel}">@{channel}</a>
      </div>
      <a class="btn btn-tg" href="https://t.me/{channel}" target="_blank">📲 Telegram kanalga o'tish</a>
      <a class="btn btn-red" href="https://t.me/{channel}" target="_blank">🛍 Buyurtma berish</a>
    </div>
  </div>

  <!-- QR -->
  <div class="qr-sec">
    <h3>📱 Mahsulot QR Kodi</h3>
    <div class="qr-img">
      <svg viewBox="0 0 200 200" xmlns="http://www.w3.org/2000/svg">
        <rect width="200" height="200" fill="white"/>
        <g fill="#1a1a2e">
          <rect x="10" y="10" width="60" height="60" rx="4"/>
          <rect x="18" y="18" width="44" height="44" rx="2" fill="white"/>
          <rect x="26" y="26" width="28" height="28" rx="1"/>
          <rect x="130" y="10" width="60" height="60" rx="4"/>
          <rect x="138" y="18" width="44" height="44" rx="2" fill="white"/>
          <rect x="146" y="26" width="28" height="28" rx="1"/>
          <rect x="10" y="130" width="60" height="60" rx="4"/>
          <rect x="18" y="138" width="44" height="44" rx="2" fill="white"/>
          <rect x="26" y="146" width="28" height="28" rx="1"/>
          <rect x="80" y="10" width="8" height="8"/><rect x="96" y="10" width="8" height="8"/><rect x="112" y="10" width="8" height="8"/>
          <rect x="80" y="26" width="8" height="8"/><rect x="104" y="26" width="8" height="8"/>
          <rect x="80" y="42" width="8" height="8"/><rect x="96" y="42" width="8" height="8"/>
          <rect x="80" y="58" width="8" height="8"/><rect x="112" y="58" width="8" height="8"/>
          <rect x="10" y="80" width="8" height="8"/><rect x="26" y="80" width="8" height="8"/>
          <rect x="80" y="80" width="8" height="8"/><rect x="104" y="80" width="8" height="8"/>
          <rect x="128" y="80" width="8" height="8"/><rect x="160" y="80" width="8" height="8"/>
          <rect x="80" y="96" width="8" height="8"/><rect x="128" y="96" width="8" height="8"/>
          <rect x="80" y="112" width="8" height="8"/><rect x="136" y="112" width="8" height="8"/>
          <rect x="80" y="128" width="8" height="8"/><rect x="104" y="128" width="8" height="8"/>
          <rect x="80" y="144" width="8" height="8"/><rect x="120" y="144" width="8" height="8"/>
          <rect x="80" y="160" width="8" height="8"/><rect x="112" y="160" width="8" height="8"/>
          <rect x="80" y="176" width="8" height="8"/><rect x="136" y="176" width="8" height="8"/>
        </g>
        <circle cx="100" cy="100" r="10" fill="#E8192C"/>
        <circle cx="100" cy="100" r="6" fill="white"/>
        <circle cx="100" cy="100" r="3" fill="#E8192C"/>
      </svg>
    </div>
    <div class="qr-code">{barcode}</div>
  </div>

</div>

<footer>© 2025 Sariosiyo Online · Barcha huquqlar himoyalangan</footer>

<!-- ADMIN LOGIN -->
<div class="overlay" id="ov" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <div class="mhandle"></div>
    <div id="s1">
      <div class="mtitle">Siz kimsiz?</div>
      <div class="msub">Sahifaga kirish uchun tanlang</div>
      <div class="role-grid">
        <div class="role-card" id="rcust" onclick="selRole('customer')">
          <div class="ri">🛍</div><div class="rn">Mijoz</div><div class="rd">Narx va ma'lumot</div>
        </div>
        <div class="role-card" id="radmin" onclick="selRole('admin')">
          <div class="ri">🔐</div><div class="rn">Admin</div><div class="rd">Stok + boshqaruv</div>
        </div>
      </div>
      <button class="mbtn mbtn-p" id="cBtn" onclick="confirmRole()" style="opacity:.4;pointer-events:none">Davom etish →</button>
      <button class="mbtn mbtn-s" onclick="closeModal()">Bekor qilish</button>
    </div>
    <div id="s2" class="pass-step">
      <div class="mtitle">🔐 Admin parol</div>
      <div class="msub">Maxfiy parolni kiriting</div>
      <div class="pass-wrap">
        <input type="password" class="pass-input" id="pi" placeholder="● ● ● ● ●" maxlength="20" onkeydown="if(event.key==='Enter')chkPass()">
        <button class="pass-toggle" onclick="tglPass()">👁</button>
      </div>
      <div class="pass-err" id="pe">❌ Noto'g'ri parol!</div>
      <div class="pass-hint">Parolni admindan oling</div>
      <button class="mbtn mbtn-p" onclick="chkPass()">✅ Kirish</button>
      <button class="mbtn mbtn-s" onclick="back1()">← Orqaga</button>
    </div>
    <div id="s3" class="pass-step">
      <div style="text-align:center;padding:10px 0 18px">
        <div style="font-size:2.8rem;margin-bottom:11px">✅</div>
        <div class="mtitle" style="color:var(--green)">Xush kelibsiz!</div>
        <div class="msub">Admin ma'lumotlari yuklandi</div>
      </div>
      <button class="mbtn mbtn-p" onclick="closeModal()">Ko'rish →</button>
    </div>
  </div>
</div>

<!-- ACTION MODAL -->
<div class="aoverlay" id="aov" onclick="if(event.target===this)closeAction()">
  <div class="amodal"><div class="mhandle"></div><div id="ac"></div></div>
</div>

<!-- CONFIRM -->
<div class="cov" id="cov">
  <div class="cbox">
    <div class="ci" id="ci">⚠️</div>
    <div class="ct" id="ct">Tasdiqlang</div>
    <div class="cm" id="cm">Davom etasizmi?</div>
    <div class="cbtns">
      <button class="cbtn cbtn-c" onclick="closeCov()">Bekor</button>
      <button class="cbtn cbtn-ok" id="cok">Ha</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
const API = '{WEBAPP_PUBLIC_URL}';
const BARCODE = '{barcode}';
const ADMIN_PWS = {json.dumps(ADMIN_PASSWORDS)};
let isAdmin = false;

// State — sahifa ochilganda bot dan kelgan ma'lumot asosida
let ST = {{
  buy: {p.buy_price},
  sell: {p.sell_price},
  stock: {p.stock},
  minStock: {p.min_stock},
  unit: '{p.unit}',
  name: `{p.name.replace('`','').replace(chr(10), ' ')}`,
  history: []
}};

// ── Admin login ──────────────────────────────────────────────────
function openModal() {{ $('ov').classList.add('open'); resetModal(); }}
function closeModal() {{ $('ov').classList.remove('open'); }}
function resetModal() {{
  $('s1').style.display='block'; $('s2').classList.remove('show'); $('s3').classList.remove('show');
  document.querySelectorAll('.role-card').forEach(c=>c.classList.remove('selected'));
  $('cBtn').style.opacity='.4'; $('cBtn').style.pointerEvents='none';
  $('pi').value=''; $('pe').classList.remove('show');
}}
function selRole(r) {{
  $('rcust').classList.toggle('selected',r==='customer');
  $('radmin').classList.toggle('selected',r==='admin');
  $('cBtn').style.opacity='1'; $('cBtn').style.pointerEvents='all';
  $('cBtn').dataset.r=r;
}}
function confirmRole() {{
  const r=$('cBtn').dataset.r;
  if(!r) return;
  if(r==='customer') {{ closeModal(); toast('✅ Mijoz sifatida kirildi'); }}
  else {{ $('s1').style.display='none'; $('s2').classList.add('show'); setTimeout(()=>$('pi').focus(),250); }}
}}
function chkPass() {{
  const p=$('pi').value.trim();
  if(ADMIN_PWS.includes(p)) {{
    $('pi').classList.remove('err'); $('pe').classList.remove('show');
    $('s2').classList.remove('show'); $('s3').classList.add('show');
    isAdmin=true; showAdmin(); setTimeout(closeModal,1300);
  }} else {{
    $('pi').classList.add('err'); $('pe').classList.add('show'); $('pi').value='';
    setTimeout(()=>$('pi').classList.remove('err'),500);
    if(navigator.vibrate) navigator.vibrate([80,40,80]);
  }}
}}
function back1() {{ $('s2').classList.remove('show'); $('s1').style.display='block'; }}
function tglPass() {{ $('pi').type=$('pi').type==='password'?'text':'password'; }}

function showAdmin() {{
  $('adminPanel').classList.add('show');
  updateStats();
  setTimeout(()=>{{}},300);
}}
function logout() {{
  isAdmin=false; $('adminPanel').classList.remove('show');
  toast('Admin paneldan chiqildi','inf');
}}
function updateStats() {{
  const profit=ST.sell-ST.buy;
  const pct=ST.buy>0?(profit/ST.buy*100).toFixed(1):0;
  $('sBuy').textContent=nf(ST.buy)+" so'm";
  $('sSell').textContent=nf(ST.sell)+" so'm";
  $('sProfit').textContent=(profit>=0?'+':'')+nf(profit)+" so'm";
  $('sStock').textContent=ST.stock+' '+ST.unit;
  $('sPct').textContent=pct+'%';
  $('pbar').style.width=Math.min(parseFloat(pct),100)+'%';
  // stok badge yangilash
  const sb=$('stockBadge');
  if(ST.stock===0){{ sb.className='stock-badge s-empty'; sb.textContent='🔴 Tugagan'; }}
  else if(ST.stock<=ST.minStock){{ sb.className='stock-badge s-low'; sb.textContent='🟡 Kam qoldi ('+ST.stock+')'; }}
  else {{ sb.className='stock-badge s-ok'; sb.textContent='🟢 Mavjud ('+ST.stock+')'; }}
}}

// ── Action modal ──────────────────────────────────────────────────
function openAction(html) {{ $('ac').innerHTML=html; $('aov').classList.add('open'); }}
function closeAction() {{ $('aov').classList.remove('open'); }}

// ── Sotish ────────────────────────────────────────────────────────
function openSotish() {{
  openAction(`
    <div class="mtitle">📤 Sotish</div>
    <div class="msub">${{ST.name}}</div>
    <div class="igroup"><label class="ilabel">Miqdor (${{ST.unit}})</label>
      <input type="number" class="finput" id="sq" value="1" min="1" max="${{ST.stock}}" oninput="calcS()"></div>
    <div class="igroup"><label class="ilabel">Narx (so'm)</label>
      <input type="number" class="finput" id="sp" value="${{ST.sell}}" oninput="calcS()"></div>
    <div class="igroup"><label class="ilabel">Izoh</label>
      <input type="text" class="finput" id="sn" placeholder="Naqd / Telegram / Boshqa"></div>
    <div class="rbox"><div class="rl">Jami summa</div><div class="rv" id="stot">${{nf(ST.sell)}} so'm</div>
      <div class="rl" style="margin-top:7px">Foyda</div>
      <div class="rv" id="sfoy">+${{nf(ST.sell-ST.buy)}} so'm</div></div>
    <div style="font-size:.8rem;color:var(--gray);margin-bottom:13px">Omborda: <b style="color:var(--white)">${{ST.stock}} ${{ST.unit}}</b></div>
    <button class="mbtn mbtn-p" onclick="doSotish()">✅ Tasdiqlash</button>
    <button class="mbtn mbtn-s" onclick="closeAction()">Bekor</button>`);
}}
function calcS() {{
  const q=+$('sq').value||1, p=+$('sp').value||ST.sell;
  $('stot').textContent=nf(q*p)+' so\'m';
  const f=q*(p-ST.buy);
  $('sfoy').textContent=(f>=0?'+':'')+nf(f)+' so\'m';
  $('sfoy').style.color=f>=0?'var(--green)':'var(--red)';
}}
async function doSotish() {{
  const q=+$('sq').value||1, p=+$('sp').value||ST.sell, n=$('sn').value||'Minisite sotish';
  if(q>ST.stock) {{ toast('❌ Stokda yetarli mahsulot yo\'q!','err'); return; }}
  const r = await apiPost('/api/sell', {{barcode:BARCODE, qty:q, price:p, note:n}});
  if(r.ok) {{
    ST.stock=r.new_stock??ST.stock-q;
    updateStats(); closeAction();
    toast('✅ '+q+' '+ST.unit+' sotildi! +'+nf(q*p)+' so\'m');
  }} else toast('❌ '+(r.error||'Xato'),'err');
}}

// ── Kirim ─────────────────────────────────────────────────────────
function openKirim() {{
  openAction(`
    <div class="mtitle">📥 Kirim</div>
    <div class="msub">Ombor kirim</div>
    <div class="igroup"><label class="ilabel">Miqdor (${{ST.unit}})</label>
      <input type="number" class="finput" id="kq" value="10" min="1" oninput="calcK()"></div>
    <div class="igroup"><label class="ilabel">Tannarx (so'm)</label>
      <input type="number" class="finput" id="kp" value="${{ST.buy}}" oninput="calcK()"></div>
    <div class="igroup"><label class="ilabel">Izoh</label>
      <input type="text" class="finput" id="kn" placeholder="Yetkazib beruvchi"></div>
    <div class="rbox warn"><div class="rl">Jami xarajat</div>
      <div class="rv y" id="ktot">${{nf(ST.buy*10)}} so'm</div>
      <div class="rl" style="margin-top:7px">Yangi stok</div>
      <div class="rv y" id="knew">${{ST.stock+10}} ${{ST.unit}}</div></div>
    <button class="mbtn mbtn-p" onclick="doKirim()">✅ Tasdiqlash</button>
    <button class="mbtn mbtn-s" onclick="closeAction()">Bekor</button>`);
}}
function calcK() {{
  const q=+$('kq').value||0, p=+$('kp').value||ST.buy;
  $('ktot').textContent=nf(q*p)+' so\'m';
  $('knew').textContent=(ST.stock+q)+' '+ST.unit;
}}
async function doKirim() {{
  const q=+$('kq').value||0, p=+$('kp').value||ST.buy, n=$('kn').value||'Minisite kirim';
  if(q<=0) {{ toast('❌ Miqdor kiriting','err'); return; }}
  const r = await apiPost('/api/supply', {{barcode:BARCODE, qty:q, buy_price:p, note:n}});
  if(r.ok) {{
    ST.stock=r.new_stock??ST.stock+q; ST.buy=p;
    updateStats(); closeAction();
    toast('✅ '+q+' '+ST.unit+' kirim qilindi!');
  }} else toast('❌ '+(r.error||'Xato'),'err');
}}

// ── Tahrirlash ────────────────────────────────────────────────────
function openTahrir() {{
  const cats=['Kiyim','Poyabzal','Aksessuar','Sport','Elektronika','Boshqa'];
  openAction(`
    <div class="mtitle">✏️ Tahrirlash</div>
    <div class="msub">Ma'lumotlarni yangilang</div>
    <div class="esect">Narxlar</div>
    <div class="igroup"><label class="ilabel">Tannarx (so'm)</label>
      <input type="number" class="finput" id="eb" value="${{ST.buy}}"></div>
    <div class="igroup"><label class="ilabel">Sotish narxi (so'm)</label>
      <input type="number" class="finput" id="es" value="${{ST.sell}}"></div>
    <div class="esect">Ma'lumotlar</div>
    <div class="igroup"><label class="ilabel">Mahsulot nomi</label>
      <input type="text" class="finput" id="en" value="${{ST.name.replace(/"/g,'&quot;')}}"></div>
    <div class="igroup"><label class="ilabel">Kategoriya</label>
      <select class="finput" id="ec">
        ${{cats.map(c=>`<option ${{c==='{p.category}'?'selected':''}}>${{c}}</option>`).join('')}}
      </select></div>
    <button class="mbtn mbtn-p" onclick="doTahrir()">💾 Saqlash</button>
    <button class="mbtn mbtn-s" onclick="closeAction()">Bekor</button>`);
}}
async function doTahrir() {{
  const fields={{buy_price:+$('eb').value, sell_price:+$('es').value,
    name:$('en').value, category:$('ec').value}};
  const r = await apiPost('/api/edit', {{barcode:BARCODE, ...fields}});
  if(r.ok) {{
    ST.buy=fields.buy_price; ST.sell=fields.sell_price; ST.name=fields.name;
    updateStats(); closeAction();
    toast('✅ Mahsulot yangilandi!');
  }} else toast('❌ '+(r.error||'Xato'),'err');
}}

// ── Tarix ─────────────────────────────────────────────────────────
async function openTarix() {{
  const r = await apiFetch('/api/history?barcode='+BARCODE);
  const items = (r?.items||[]).map(h=>`
    <div class="hist-item">
      <div class="hi-l">
        <div style="font-weight:700">${{h.type==='sell'?'📤 Sotish':'📥 Kirim'}} — ${{h.qty}} ${{ST.unit}}</div>
        <div class="hi-d">${{h.date}} · ${{h.note}}</div>
      </div>
      <div class="hi-a ${{h.type==='sell'?'g':'r'}}">${{h.type==='sell'?'+':'-'}}${{nf(h.total)}} so'm</div>
    </div>`).join('');
  openAction(`
    <div class="mtitle">📊 Operatsiya tarixi</div>
    <div class="msub">${{r?.items?.length||0}} ta operatsiya</div>
    ${{items||'<div style="color:var(--gray);text-align:center;padding:20px">Tarix yo\'q</div>'}}
    <button class="mbtn mbtn-s" style="margin-top:10px" onclick="closeAction()">Yopish</button>`);
}}

// ── Confirm archive ───────────────────────────────────────────────
function confirmArchive() {{
  showCov('🗑','Arxivlash','Mahsulot ro\'yxatdan o\'chiriladi. Tasdiqlaysizmi?', async ()=>{{
    const r = await apiPost('/api/archive',{{barcode:BARCODE}});
    closeCov(); closeAction();
    toast('✅ Arxivlandi');
    $('adminPanel').classList.remove('show');
    document.querySelector('.card').style.opacity='.4';
    document.querySelector('.card').style.filter='grayscale(.7)';
  }});
}}
function copyLink() {{
  navigator.clipboard?.writeText(window.location.href).then(()=>toast('🔗 Link nusxalandi','inf'));
}}

// ── Confirm dialog ────────────────────────────────────────────────
function showCov(icon,title,msg,cb) {{
  $('ci').textContent=icon; $('ct').textContent=title; $('cm').textContent=msg;
  $('cok').onclick=cb; $('cov').classList.add('open');
}}
function closeCov() {{ $('cov').classList.remove('open'); }}

// ── API ───────────────────────────────────────────────────────────
async function apiPost(path, data) {{
  try {{
    const r = await fetch(API+path,{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(data)}});
    return await r.json();
  }} catch(e) {{ return {{ok:false,error:e.message}}; }}
}}
async function apiFetch(path) {{
  try {{
    const r = await fetch(API+path);
    return await r.json();
  }} catch(e) {{ return null; }}
}}

// ── Helpers ───────────────────────────────────────────────────────
function $(id) {{ return document.getElementById(id); }}
function nf(n) {{ return Math.round(n||0).toLocaleString('uz-UZ'); }}
function toast(msg,type='',dur=2500) {{
  const t=$('toast'); t.textContent=msg;
  t.className='toast show'+(type?' '+type:'');
  setTimeout(()=>t.classList.remove('show'),dur);
}}
</script>
</body>
</html>"""

def get_scanner_html() -> str:
    """Skaner sahifasi — to'liq kamera + jsQR + bot API"""
    admin_pws_json = json.dumps(ADMIN_PASSWORDS)
    return f"""<!DOCTYPE html>
<html lang="uz">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,user-scalable=no">
<title>Sariosiyo — QR Skaner</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700;800&display=swap');
:root{{--red:#E8192C;--red2:#C0141F;--dark:#0A0A0F;--card:#111118;--card2:#18181F;
  --border:rgba(255,255,255,.08);--white:#F0F2F8;--gray:#7A8499;
  --green:#1ED760;--yellow:#F5A623;--blue:#4A9EFF}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Space Grotesk',sans-serif;background:var(--dark);color:var(--white);min-height:100vh}}
.header{{background:var(--red);padding:0 18px;height:56px;display:flex;align-items:center;
  justify-content:space-between;position:sticky;top:0;z-index:100}}
.hl{{display:flex;align-items:center;gap:10px}}
.logo{{width:36px;height:36px;background:rgba(255,255,255,.2);border-radius:10px;
  display:flex;align-items:center;justify-content:center;font-size:18px}}
.ht{{font-weight:800;font-size:1rem;letter-spacing:.5px}}
.hs{{font-size:.68rem;opacity:.85}}
.hbadge{{background:rgba(255,255,255,.18);border:1px solid rgba(255,255,255,.3);color:white;
  font-size:.72rem;font-weight:600;padding:5px 12px;border-radius:20px}}
.tabs{{display:flex;background:var(--card);border-bottom:1px solid var(--border)}}
.tab{{flex:1;padding:13px 8px;text-align:center;font-size:.78rem;font-weight:600;
  cursor:pointer;color:var(--gray);border-bottom:2px solid transparent;transition:all .2s}}
.tab.active{{color:var(--white);border-bottom-color:var(--red)}}
.pane{{display:none;flex-direction:column;align-items:center;padding:16px;gap:14px}}
.pane.active{{display:flex}}
.camwrap{{position:relative;width:100%;max-width:400px;background:#000;border-radius:20px;
  overflow:hidden;aspect-ratio:1/1;box-shadow:0 0 0 1px var(--border),0 20px 60px rgba(0,0,0,.6)}}
#video{{width:100%;height:100%;object-fit:cover;display:block}}
#canvas{{display:none}}
.sovl{{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;pointer-events:none}}
.sbox{{width:65%;height:65%;position:relative}}
.sline{{position:absolute;left:5%;right:5%;height:2px;
  background:linear-gradient(90deg,transparent,var(--red),transparent);
  box-shadow:0 0 8px rgba(232,25,44,.8);animation:mv 2s ease-in-out infinite}}
@keyframes mv{{0%{{top:8%}}50%{{top:88%}}100%{{top:8%}}}}
.sc{{position:absolute;width:22px;height:22px;border-color:var(--red);border-style:solid}}
.tl{{top:0;left:0;border-width:3px 0 0 3px;border-radius:4px 0 0 0}}
.tr{{top:0;right:0;border-width:3px 3px 0 0;border-radius:0 4px 0 0}}
.bl{{bottom:0;left:0;border-width:0 0 3px 3px;border-radius:0 0 0 4px}}
.br{{bottom:0;right:0;border-width:0 3px 3px 0;border-radius:0 0 4px 0}}
.cst{{position:absolute;bottom:0;left:0;right:0;background:linear-gradient(transparent,rgba(0,0,0,.85));
  padding:20px 14px 12px;font-size:.78rem;color:rgba(255,255,255,.8);text-align:center}}
.camwrap.found{{animation:fg .4s ease}}
@keyframes fg{{0%,100%{{box-shadow:0 0 0 1px var(--border)}}50%{{box-shadow:0 0 0 3px var(--green),0 0 30px rgba(30,215,96,.4)}}}}
.pills{{display:flex;flex-wrap:wrap;gap:6px;width:100%;max-width:400px}}
.pill{{padding:5px 12px;border-radius:20px;font-size:.72rem;font-weight:600;
  background:rgba(255,255,255,.06);color:var(--gray);border:1px solid var(--border);cursor:pointer;transition:all .2s}}
.pill.on{{background:rgba(232,25,44,.18);color:var(--white);border-color:var(--red)}}
.stbar{{width:100%;max-width:400px;padding:10px 16px;border-radius:12px;font-size:.83rem;
  font-weight:500;text-align:center;background:var(--card2);border:1px solid var(--border);transition:all .3s}}
.stbar.scanning{{color:var(--yellow);border-color:rgba(245,166,35,.3)}}
.stbar.found{{color:var(--green);border-color:rgba(30,215,96,.3)}}
.stbar.error{{color:var(--red);border-color:rgba(232,25,44,.3)}}
.loader{{width:100%;max-width:400px;display:none;align-items:center;justify-content:center;
  gap:12px;padding:22px;background:var(--card2);border-radius:18px;border:1px solid var(--border)}}
.loader.show{{display:flex}}
.spin{{width:26px;height:26px;border:2.5px solid rgba(255,255,255,.1);border-top-color:var(--red);
  border-radius:50%;animation:sp .7s linear infinite}}
@keyframes sp{{to{{transform:rotate(360deg)}}}}
.rcard{{width:100%;max-width:400px;background:var(--card2);border-radius:20px;
  border:1px solid var(--border);overflow:hidden;display:none;
  animation:su .35s cubic-bezier(.16,1,.3,1)}}
.rcard.show{{display:block}}
@keyframes su{{from{{opacity:0;transform:translateY(24px)}}to{{opacity:1;transform:translateY(0)}}}}
.ri{{width:100%;height:175px;object-fit:cover;display:block}}
.riph{{width:100%;height:130px;background:linear-gradient(135deg,#0d0d14,#1a1a2e);
  display:flex;align-items:center;justify-content:center;font-size:3.5rem}}
.rbody{{padding:15px}}
.rbadge{{display:inline-block;background:var(--red);color:white;font-size:.67rem;font-weight:700;
  padding:3px 9px;border-radius:18px;margin-bottom:7px}}
.rname{{font-size:1.1rem;font-weight:700;line-height:1.3;margin-bottom:3px}}
.rcode{{font-size:.73rem;color:var(--gray);font-family:monospace;margin-bottom:10px}}
.rprice{{font-size:1.6rem;font-weight:800;color:var(--green);margin-bottom:9px}}
.rprice span{{font-size:.83rem;color:var(--gray);font-weight:400}}
.rstk{{display:inline-flex;align-items:center;gap:5px;padding:5px 12px;border-radius:18px;
  font-size:.78rem;font-weight:600;margin-bottom:10px}}
.so{{background:rgba(30,215,96,.15);color:var(--green)}}
.sl{{background:rgba(245,166,35,.15);color:var(--yellow)}}
.se{{background:rgba(232,25,44,.15);color:var(--red)}}
.rgrid{{display:grid;grid-template-columns:1fr 1fr;gap:7px;margin-bottom:10px}}
.rinfo{{background:rgba(255,255,255,.04);border-radius:9px;padding:8px 10px}}
.ril{{font-size:.63rem;color:var(--gray);text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px}}
.riv{{font-size:.86rem;font-weight:600}}
.radm{{background:rgba(232,25,44,.08);border:1px solid rgba(232,25,44,.2);
  border-radius:9px;padding:9px 11px;margin-bottom:10px;font-size:.79rem;line-height:1.8;color:var(--gray)}}
.rdiv{{height:1px;background:var(--border);margin:9px 0}}
.rcon{{font-size:.83rem;color:var(--gray);line-height:2}}
.rcon a{{color:var(--white);text-decoration:none}}
.racts{{display:flex;gap:7px;margin-top:10px}}
.rbtn{{flex:1;padding:10px 6px;border-radius:11px;border:none;font-family:inherit;
  font-size:.8rem;font-weight:700;cursor:pointer;text-decoration:none;
  display:flex;align-items:center;justify-content:center;gap:4px;transition:all .2s}}
.rbtn:active{{opacity:.8;transform:scale(.97)}}
.btg{{background:#2481CC;color:white}}
.bred{{background:var(--red);color:white}}
.bscan{{background:transparent;border:1.5px solid var(--red);color:var(--red)}}
.nf{{width:100%;max-width:400px;background:var(--card2);border-radius:20px;
  border:1px solid rgba(232,25,44,.2);padding:28px 18px;text-align:center;display:none}}
.nf.show{{display:block;animation:su .3s ease}}
.nfi{{font-size:2.8rem;margin-bottom:9px}}
.nft{{font-size:.97rem;font-weight:700;color:var(--red);margin-bottom:4px}}
.nfs{{font-size:.8rem;color:var(--gray);margin-bottom:14px}}
.mrow{{width:100%;max-width:400px;display:flex;gap:8px}}
.minput{{flex:1;background:var(--card2);border:1.5px solid var(--border);color:var(--white);
  font-family:inherit;font-size:.93rem;padding:12px 13px;border-radius:12px;outline:none;transition:border-color .2s}}
.minput:focus{{border-color:var(--red)}}
.mbtn{{background:var(--red);color:white;border:none;padding:12px 17px;border-radius:12px;
  font-family:inherit;font-size:.93rem;font-weight:600;cursor:pointer;white-space:nowrap}}
.mbtn:active{{opacity:.85}}
.hl2{{width:100%;max-width:400px;display:flex;justify-content:space-between;align-items:center}}
.hlist{{width:100%;max-width:400px}}
.hitem{{background:var(--card2);border-radius:12px;padding:11px 13px;margin-bottom:7px;
  border:1px solid var(--border);display:flex;align-items:center;gap:9px;cursor:pointer;transition:border-color .2s}}
.hitem:hover{{border-color:rgba(232,25,44,.3)}}
.hcode{{font-family:monospace;font-size:.8rem;color:var(--gray);flex:1}}
.hname{{font-size:.86rem;font-weight:600}}
.htime{{font-size:.7rem;color:var(--gray);margin-top:1px}}
.hdel{{background:none;border:none;color:var(--gray);cursor:pointer;font-size:.95rem;padding:4px}}
.hdel:hover{{color:var(--red)}}
.hemp{{color:var(--gray);font-size:.86rem;text-align:center;padding:28px}}
.sw{{width:100%;max-width:400px}}
.slbl{{font-size:.7rem;color:var(--gray);text-transform:uppercase;letter-spacing:.5px;margin-bottom:7px;padding-left:2px}}
.srow{{background:var(--card2);border-radius:12px;padding:13px;margin-bottom:7px;
  display:flex;align-items:center;justify-content:space-between;border:1px solid var(--border)}}
.st{{font-size:.88rem;font-weight:600}}
.ss{{font-size:.72rem;color:var(--gray);margin-top:2px}}
.tgl{{position:relative;width:44px;height:24px;cursor:pointer}}
.tgl input{{opacity:0;width:0;height:0;position:absolute}}
.sldr{{position:absolute;cursor:pointer;inset:0;background:rgba(255,255,255,.15);border-radius:24px;transition:.3s}}
.sldr:before{{content:"";position:absolute;height:18px;width:18px;left:3px;bottom:3px;
  background:white;border-radius:50%;transition:.3s}}
input:checked+.sldr{{background:var(--red)}}
input:checked+.sldr:before{{transform:translateX(20px)}}
.admpass{{display:none;background:rgba(232,25,44,.08);border:1px solid rgba(232,25,44,.2);
  border-radius:9px;padding:10px 12px;margin-top:6px}}
.admpass.show{{display:block}}
.toast{{position:fixed;bottom:26px;left:50%;transform:translateX(-50%) translateY(100px);
  background:#22C55E;color:white;padding:10px 20px;border-radius:20px;
  font-size:.86rem;font-weight:600;z-index:999;
  transition:transform .3s cubic-bezier(.16,1,.3,1);white-space:nowrap;
  box-shadow:0 4px 20px rgba(0,0,0,.4)}}
.toast.show{{transform:translateX(-50%) translateY(0)}}
.toast.err{{background:var(--red)}}.toast.inf{{background:var(--blue)}}
.pb{{height:22px}}
</style>
</head>
<body>

<div class="header">
  <div class="hl">
    <div class="logo">🏪</div>
    <div><div class="ht">Sariosiyo</div><div class="hs">QR & Barcode Skaner</div></div>
  </div>
  <div class="hbadge" id="srv">⏳ Ulanmoqda</div>
</div>

<div class="tabs">
  <div class="tab active" onclick="sw('scan')">📷 Skanerlash</div>
  <div class="tab" onclick="sw('manual')">⌨️ Qo'lda</div>
  <div class="tab" onclick="sw('hist')">🕘 Tarix</div>
  <div class="tab" onclick="sw('sett')">⚙️ Sozlama</div>
</div>

<!-- SCAN -->
<div class="pane active" id="pane-scan">
  <div class="camwrap" id="cw">
    <video id="video" autoplay playsinline muted></video>
    <canvas id="canvas"></canvas>
    <div class="sovl"><div class="sbox"><div class="sline"></div>
      <div class="sc tl"></div><div class="sc tr"></div>
      <div class="sc bl"></div><div class="sc br"></div></div></div>
    <div class="cst" id="cst">📷 Kamera yuklanmoqda...</div>
  </div>
  <div class="pills" id="pills"></div>
  <div class="stbar" id="stbar">Kamera ochilishini kuting...</div>
  <div class="loader" id="ld"><div class="spin"></div><span style="font-size:.86rem;color:var(--gray)">Qidirilmoqda...</span></div>
  <div class="rcard" id="rc"></div>
  <div class="nf" id="nf">
    <div class="nfi">🔍</div>
    <div class="nft">Mahsulot topilmadi</div>
    <div class="nfs" id="nfs">Barcode bazada yo'q</div>
    <button onclick="reset()" class="rbtn bscan" style="max-width:160px;margin:0 auto">↩ Qayta skanerlash</button>
  </div>
  <div class="pb"></div>
</div>

<!-- MANUAL -->
<div class="pane" id="pane-manual">
  <div class="mrow">
    <input class="minput" id="mi" type="text" placeholder="SAR-123456 yoki mahsulot kodi..."
      autocomplete="off" autocorrect="off" autocapitalize="characters">
    <button class="mbtn" onclick="searchM()">🔍</button>
  </div>
  <div class="loader" id="ld2"><div class="spin"></div><span style="font-size:.86rem;color:var(--gray)">Qidirilmoqda...</span></div>
  <div class="rcard" id="rc2"></div>
  <div class="nf" id="nf2"><div class="nfi">🔍</div><div class="nft">Topilmadi</div><div class="nfs" id="nfs2"></div></div>
  <div class="pb"></div>
</div>

<!-- HISTORY -->
<div class="pane" id="pane-hist">
  <div class="hl2">
    <span style="font-weight:700;font-size:.93rem">Oxirgi skanlar</span>
    <button onclick="clearH()" style="background:none;border:1px solid var(--red);color:var(--red);border-radius:8px;padding:4px 10px;cursor:pointer;font-size:.76rem">Tozalash</button>
  </div>
  <div class="hlist" id="hl"></div>
  <div class="pb"></div>
</div>

<!-- SETTINGS -->
<div class="pane" id="pane-sett">
  <div class="sw">
    <div class="slbl" style="margin-top:.5rem">Server</div>
    <div class="srow">
      <div><div class="st">Bot Server</div><div class="ss" id="surl">{WEBAPP_PUBLIC_URL}</div></div>
      <span id="sconn" style="font-size:.78rem;color:var(--gray)">⏳</span>
    </div>
    <div class="slbl" style="margin-top:1rem">Kamera</div>
    <div class="srow">
      <div><div class="st">Orqa kamera</div><div class="ss">Barcode uchun qulay</div></div>
      <label class="tgl"><input type="checkbox" id="backCam" checked onchange="restartCam()"><span class="sldr"></span></label>
    </div>
    <div class="srow">
      <div><div class="st">Torch</div><div class="ss">Qorong'i joyda</div></div>
      <label class="tgl"><input type="checkbox" id="torch" onchange="tgTorch()"><span class="sldr"></span></label>
    </div>
    <div class="slbl" style="margin-top:1rem">Signal</div>
    <div class="srow">
      <div><div class="st">Beep ovozi</div><div class="ss">Topilganda signal</div></div>
      <label class="tgl"><input type="checkbox" id="snd" checked><span class="sldr"></span></label>
    </div>
    <div class="srow">
      <div><div class="st">Tebranish</div><div class="ss">Topilganda vibro</div></div>
      <label class="tgl"><input type="checkbox" id="vib" checked><span class="sldr"></span></label>
    </div>
    <div class="slbl" style="margin-top:1rem">Admin</div>
    <div class="srow">
      <div><div class="st">Admin rejimi</div><div class="ss" id="admLbl">Stok va narxlarni ko'rish</div></div>
      <label class="tgl"><input type="checkbox" id="admMode" onchange="tgAdm()"><span class="sldr"></span></label>
    </div>
    <div class="admpass" id="apw">
      <div style="font-size:.76rem;margin-bottom:5px;color:var(--gray)">Admin paroli:</div>
      <div style="display:flex;gap:6px">
        <input type="password" id="api" placeholder="Parol..." class="minput" style="font-size:.83rem;padding:8px 12px">
        <button onclick="chkAdm()" class="mbtn" style="padding:8px 13px;font-size:.83rem">✓</button>
      </div>
    </div>
  </div>
  <div class="pb"></div>
</div>

<div class="toast" id="t"></div>

<script src="https://cdn.jsdelivr.net/npm/jsqr@1.4.0/dist/jsQR.js"></script>
<script>
const API='{WEBAPP_PUBLIC_URL}';
const APWS={admin_pws_json};
let scanning=true,lastCode='',lastTime=0;
let camStream=null,torchTrack=null,raf=null;
let isAdm=false,srvOk=false;

const FMTS=[
  {{id:'qr_code',label:'QR',on:true}},
  {{id:'code_128',label:'CODE 128',on:true}},
  {{id:'code_39',label:'CODE 39',on:true}},
  {{id:'ean_13',label:'EAN-13',on:true}},
  {{id:'ean_8',label:'EAN-8',on:false}},
  {{id:'upc_a',label:'UPC-A',on:false}},
];

const $=id=>document.getElementById(id);

// Pilllar
function buildPills(){{
  const w=$('pills'); w.innerHTML='';
  FMTS.forEach(f=>{{
    const d=document.createElement('div');
    d.className='pill'+(f.on?' on':'');
    d.textContent=f.label;
    d.onclick=()=>{{f.on=!f.on;d.classList.toggle('on',f.on);}};
    w.appendChild(d);
  }});
}}
buildPills();

// Server check
async function chkSrv(){{
  try{{
    const r=await fetch(API+'/health',{{signal:AbortSignal.timeout(5000)}});
    srvOk=r.ok;
  }}catch{{srvOk=false;}}
  const b=$('srv'),c=$('sconn');
  if(srvOk){{
    b.textContent='🟢 Online';b.style.background='rgba(30,215,96,.2)';b.style.borderColor='rgba(30,215,96,.4)';
    c.textContent='🟢 Ulangan';c.style.color='var(--green)';
  }}else{{
    b.textContent='🔴 Offline';b.style.background='rgba(232,25,44,.2)';b.style.borderColor='rgba(232,25,44,.4)';
    c.textContent='🔴 Ulanmagan';c.style.color='var(--red)';
  }}
}}

// Tabs
function sw(name){{
  document.querySelectorAll('.tab').forEach((t,i)=>t.classList.toggle('active',['scan','manual','hist','sett'][i]===name));
  document.querySelectorAll('.pane').forEach(p=>p.classList.remove('active'));
  $('pane-'+name).classList.add('active');
  if(name==='hist') renderH();
  scanning=(name==='scan');
  if(name==='scan'&&raf){{cancelAnimationFrame(raf);scanLoop();}}
}}

// Camera
async function startCam(){{
  const back=$('backCam').checked;
  try{{
    if(camStream) camStream.getTracks().forEach(t=>t.stop());
    camStream=await navigator.mediaDevices.getUserMedia({{video:{{facingMode:back?'environment':'user',width:{{ideal:1280}},height:{{ideal:720}}}}}});
    const vid=$('video');
    vid.srcObject=camStream;
    torchTrack=camStream.getVideoTracks()[0];
    vid.onloadedmetadata=()=>{{vid.play();$('cst').textContent='';setSt('🟡 QR yoki barcode ko\'rsating','scanning');}};
  }}catch(e){{
    $('cw').style.display='none';
    setSt('⌨️ Kamera yo\'q — qo\'lda kiriting','error');
  }}
}}
async function restartCam(){{scanning=true;if(raf)cancelAnimationFrame(raf);await startCam();scanLoop();}}
async function tgTorch(){{
  if(!torchTrack)return;
  try{{await torchTrack.applyConstraints({{advanced:[{{torch:$('torch').checked}}]}});}}
  catch{{$('torch').checked=false;toast('Torch qo\'llab-quvvatlanmaydi','err');}}
}}

// Scan loop
function scanLoop(){{
  const vid=$('video'),cv=$('canvas'),ctx=cv.getContext('2d',{{willReadFrequently:true}});
  const useNative='BarcodeDetector' in window;
  let det=null;
  async function loop(){{
    if(!scanning)return;
    if(vid.readyState===vid.HAVE_ENOUGH_DATA){{
      cv.width=vid.videoWidth;cv.height=vid.videoHeight;
      ctx.drawImage(vid,0,0,cv.width,cv.height);
      let code=null;
      // 1. BarcodeDetector (Chrome)
      if(useNative){{
        try{{
          const fmts=FMTS.filter(f=>f.on).map(f=>f.id);
          if(!det||det._f!==fmts.join()) {{det=new BarcodeDetector({{formats:fmts}});det._f=fmts.join();}}
          const res=await det.detect(vid);
          if(res.length>0) code=res[0].rawValue;
        }}catch{{det=null;}}
      }}
      // 2. jsQR fallback (iOS Safari, Firefox)
      if(!code&&window.jsQR){{
        try{{
          const id=ctx.getImageData(0,0,cv.width,cv.height);
          const qr=jsQR(id.data,id.width,id.height,{{inversionAttempts:'dontInvert'}});
          if(qr) code=qr.data;
        }}catch{{}}
      }}
      if(code){{
        const now=Date.now();
        if(code!==lastCode||now-lastTime>3000){{
          lastCode=code;lastTime=now;
          beep();vib();
          $('cw').classList.add('found');
          setTimeout(()=>$('cw').classList.remove('found'),600);
          scanning=false;
          await lookup(code,'rc','nf','nfs','ld');
          return;
        }}
      }}
    }}
    if(scanning) raf=requestAnimationFrame(loop);
  }}
  vid.addEventListener('playing',()=>{{raf=requestAnimationFrame(loop);}},{{once:false}});
  raf=requestAnimationFrame(loop);
}}

// URL barcode extraction
function extractBarcode(raw){{
  try{{
    const u=new URL(raw);
    const parts=u.pathname.split('/').filter(Boolean);
    if(parts[0]==='p'&&parts[1]) return parts[1];
    const start=u.searchParams.get('start')||'';
    if(start.startsWith('qr_')) return start.slice(3);
    if(u.searchParams.get('barcode')) return u.searchParams.get('barcode');
  }}catch{{}}
  return raw;
}}

// Lookup
async function lookup(raw,cid,nid,nsid,lid){{
  showEl(lid,true);hide(cid);hide(nid);
  const code=extractBarcode(raw);
  setSt('⏳ '+code,'scanning');
  addH(code);
  const data=await apiFetch('/api/product?barcode='+encodeURIComponent(code));
  showEl(lid,false);
  if(data&&data.found){{
    renderCard(data.product,code,cid);
    setSt('✅ Topildi: '+data.product.name,'found');
    show(cid);
  }}else if(data&&!data.found){{
    $(nsid).textContent='Barcode: '+code;
    show(nid);setSt('❌ Topilmadi: '+code,'error');
  }}else{{
    $(nsid).textContent='Server bilan aloqa yo\'q. Kod: '+code;
    show(nid);setSt('❌ Server offline','error');
  }}
}}

// Render card
function renderCard(p,raw,cid){{
  const sc=p.stock===0?'se':(p.stock<=(p.min_stock||5)?'sl':'so');
  const st=p.stock===0?'🔴 Tugagan':(p.stock<=(p.min_stock||5)?`🟡 Kam: ${{p.stock}} ${{p.unit}}`:`🟢 Mavjud: ${{p.stock}} ${{p.unit}}`);
  const img=p.photo_url
    ?`<img class="ri" src="${{p.photo_url}}" onerror="this.outerHTML='<div class=riph>📦</div>'">`
    :`<div class="riph">📦</div>`;
  let grid='';
  if(p.color) grid+=`<div class="rinfo"><div class="ril">Rang</div><div class="riv">${{p.color}}</div></div>`;
  if(p.size)  grid+=`<div class="rinfo"><div class="ril">O'lcham</div><div class="riv">${{p.size}}</div></div>`;
  if(p.category) grid+=`<div class="rinfo"><div class="ril">Kategoriya</div><div class="riv">${{p.category}}</div></div>`;
  grid+=`<div class="rinfo"><div class="ril">Birlik</div><div class="riv">${{p.unit}}</div></div>`;
  let admBlk='';
  if(isAdm){{
    const pr=(p.sell_price||0)-(p.buy_price||0);
    const pct=p.buy_price>0?((pr/p.buy_price)*100).toFixed(0):0;
    admBlk=`<div class="radm">
      💰 Tannarx: <b style="color:var(--white)">${{nf(p.buy_price)}} so'm</b><br>
      📈 Foyda: <b style="color:var(--green)">+${{nf(pr)}} so'm (${{pct}}%)</b><br>
      📦 Stok: <b style="color:var(--white)">${{p.stock}} ${{p.unit}}</b></div>`;
  }}
  const pubUrl=`${{API}}/p/${{encodeURIComponent(p.barcode||raw)}}`;
  $(cid).innerHTML=`
    ${{img}}
    <div class="rbody">
      <div class="rbadge">${{p.category||'Mahsulot'}}</div>
      <div class="rname">${{p.name}}</div>
      <div class="rcode">${{p.code}} · ${{p.barcode||raw}}</div>
      <div class="rprice">${{nf(p.sell_price)}} <span>so'm</span></div>
      <div class="rstk ${{sc}}">${{st}}</div>
      ${{admBlk}}
      ${{grid?`<div class="rgrid">${{grid}}</div>`:''}}
      ${{p.description?`<div style="font-size:.8rem;color:var(--gray);line-height:1.6;padding:8px;background:rgba(255,255,255,.03);border-radius:8px;margin-bottom:9px">${{p.description}}</div>`:''}}
      <div class="rdiv"></div>
      <div class="rcon">📱 <a href="tel:+998909182186">+998 90 918 21 86</a><br>📱 <a href="tel:+998947043111">+998 94 704 31 11</a></div>
      <div class="racts">
        <a href="https://t.me/{CHANNEL_ID.lstrip('@')}" target="_blank" class="rbtn btg">📲 Telegram</a>
        <a href="${{pubUrl}}" target="_blank" class="rbtn bred">🌐 Sahifa</a>
        <button onclick="reset()" class="rbtn bscan">↩ Qayta</button>
      </div>
    </div>`;
}}

// Manual search
async function searchM(){{
  const code=$('mi').value.trim();
  if(!code){{$('mi').focus();return;}}
  beep();
  await lookup(code,'rc2','nf2','nfs2','ld2');
}}
$('mi').addEventListener('keydown',e=>{{if(e.key==='Enter')searchM();}});

// Reset
function reset(){{
  scanning=true;lastCode='';
  hide('rc');hide('nf');hide('rc2');hide('nf2');
  setSt('🟡 QR yoki barcode ko\'rsating','scanning');
  if(raf) cancelAnimationFrame(raf);
  scanLoop();
}}

// History
function addH(code){{
  let h=JSON.parse(localStorage.getItem('sarH')||'[]');
  h=h.filter(x=>x.code!==code);
  h.unshift({{code,time:Date.now()}});
  localStorage.setItem('sarH',JSON.stringify(h.slice(0,40)));
}}
function renderH(){{
  const list=$('hl');
  const h=JSON.parse(localStorage.getItem('sarH')||'[]');
  if(!h.length){{list.innerHTML='<div class="hemp">Hali hech narsa skanlanmagan</div>';return;}}
  list.innerHTML=h.map((item,i)=>`
    <div class="hitem" onclick="histLook('${{item.code}}')">
      <span style="font-size:1.1rem">📊</span>
      <div style="flex:1"><div class="hname">${{item.code}}</div>
        <div class="htime">${{new Date(item.time).toLocaleString('uz-UZ')}}</div></div>
      <button class="hdel" onclick="event.stopPropagation();rmH(${{i}})">✕</button>
    </div>`).join('');
}}
async function histLook(code){{sw('scan');await lookup(code,'rc','nf','nfs','ld');}}
function rmH(i){{let h=JSON.parse(localStorage.getItem('sarH')||'[]');h.splice(i,1);localStorage.setItem('sarH',JSON.stringify(h));renderH();}}
function clearH(){{localStorage.removeItem('sarH');renderH();toast('Tarix tozalandi','inf');}}

// Admin
function tgAdm(){{
  const on=$('admMode').checked;
  if(on){{$('apw').classList.add('show');$('admLbl').textContent='Parol kiriting';$('api').focus();}}
  else{{isAdm=false;$('apw').classList.remove('show');$('admLbl').textContent='Stok va narxlarni ko\'rish';toast('Admin rejimdan chiqildi','inf');}}
}}
function chkAdm(){{
  const p=$('api').value.trim();
  if(APWS.includes(p)){{
    isAdm=true;$('apw').classList.remove('show');
    $('admLbl').textContent='✅ Admin rejim faol';
    toast('✅ Admin rejim yoqildi');
  }}else{{
    $('api').value='';$('api').style.borderColor='var(--red)';
    setTimeout(()=>$('api').style.borderColor='',1000);
    toast('❌ Noto\'g\'ri parol','err');
    $('admMode').checked=false;tgAdm();
  }}
}}

// API
async function apiFetch(path){{
  try{{const r=await fetch(API+path);return await r.json();}}
  catch{{return null;}}
}}

// Helpers
function setSt(t,c){{const e=$('stbar');if(e){{e.textContent=t;e.className='stbar '+(c||'');}}}}
function show(id){{const e=$(id);if(e)e.classList.add('show');}}
function hide(id){{const e=$(id);if(e)e.classList.remove('show');}}
function showEl(id,v){{const e=$(id);if(e)e.classList.toggle('show',v);}}
function toast(msg,type='',dur=2500){{
  const t=$('t');t.textContent=msg;t.className='toast show'+(type?' '+type:'');
  setTimeout(()=>t.classList.remove('show'),dur);
}}
function beep(){{
  if(!$('snd').checked)return;
  try{{const ctx=new(window.AudioContext||window.webkitAudioContext)();
    const o=ctx.createOscillator(),g=ctx.createGain();
    o.connect(g);g.connect(ctx.destination);
    o.type='square';o.frequency.value=900;
    g.gain.setValueAtTime(.25,ctx.currentTime);
    g.gain.exponentialRampToValueAtTime(.001,ctx.currentTime+.1);
    o.start(ctx.currentTime);o.stop(ctx.currentTime+.1);}}catch{{}}
}}
function vib(){{if($('vib').checked&&navigator.vibrate)navigator.vibrate(80);}}
function nf(n){{return Math.round(n||0).toLocaleString('uz-UZ');}}

// Init
chkSrv();setInterval(chkSrv,30000);
startCam().then(()=>scanLoop());
</script>
</body>
</html>"""

# ════════════════════════════════════════════════════════════════
#  5. HTTP SERVER — barcha endpointlar
# ════════════════════════════════════════════════════════════════
async def start_http_server():
    try:
        from aiohttp import web
        import aiohttp_cors
    except ImportError:
        logger.warning("pip install aiohttp aiohttp-cors")
        return

    async def _json(data, status=200):
        from aiohttp import web
        return web.Response(
            text=json.dumps(data, ensure_ascii=False),
            content_type="application/json",
            status=status
        )

    # GET /health
    async def health(req):
        from aiohttp import web
        return web.Response(text="OK")

    # GET /scanner
    async def scanner(req):
        from aiohttp import web
        return web.Response(text=get_scanner_html(), content_type="text/html", charset="utf-8")

    # GET /p/{barcode} — public mahsulot sahifasi
    async def public_page(req):
        from aiohttp import web
        barcode = req.match_info.get("barcode","").strip().upper() or \
                  req.rel_url.query.get("barcode","").strip().upper()
        if not barcode:
            return web.Response(text="<h2>Barcode kiritilmadi</h2>", content_type="text/html")
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
        if not p:
            html = f"""<!DOCTYPE html><html lang="uz"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Topilmadi</title>
<style>body{{font-family:sans-serif;background:#0D0D14;color:#F4F6FB;
  display:flex;align-items:center;justify-content:center;min-height:100vh;flex-direction:column;gap:14px}}
a{{color:#E8192C}}</style></head>
<body><div style="font-size:3rem">🔍</div>
<div style="font-size:1.1rem;color:#E8192C;font-weight:700">Mahsulot topilmadi</div>
<div style="color:#8892A4;font-size:.88rem">Barcode: {barcode}</div>
<a href="https://t.me/{CHANNEL_ID.lstrip('@')}">@{CHANNEL_ID.lstrip('@')}</a>
</body></html>"""
            return web.Response(text=html, content_type="text/html", charset="utf-8")
        return web.Response(text=get_public_page_html(p), content_type="text/html", charset="utf-8")

    # GET /api/product?barcode=
    async def api_product(req):
        barcode = req.rel_url.query.get("barcode","").strip().upper()
        if not barcode:
            return await _json({"found":False,"error":"barcode required"}, 400)
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
            if not p:
                return await _json({"found":False})
            # oylik statistika
            now = datetime.now()
            ms = await s.execute(
                select(func.sum(Sale.quantity), func.sum(Sale.sell_price*Sale.quantity))
                .where(Sale.product_id==p.id,
                       extract("year",Sale.sold_at)==now.year,
                       extract("month",Sale.sold_at)==now.month)
            )
            mr = ms.one()
        return await _json({"found":True,"product":{
            "id":p.id,"code":p.code,"barcode":p.barcode,
            "name":p.name,"category":p.category,"description":p.description,
            "color":p.color,"size":p.size,"unit":p.unit,
            "buy_price":p.buy_price,"sell_price":p.sell_price,
            "stock":p.stock,"min_stock":p.min_stock,
            "photo_url":p.photo_url or "",
            "monthly_sold":mr[0] or 0,"monthly_revenue":mr[1] or 0,
        }})

    # POST /api/sell
    async def api_sell(req):
        try: data = await req.json()
        except: return await _json({"ok":False,"error":"JSON xato"}, 400)
        barcode   = str(data.get("barcode","")).strip().upper()
        qty       = int(data.get("qty",1))
        price     = float(data.get("price",0)) or None
        note      = str(data.get("note","Minisite"))
        api_token = str(data.get("token",""))
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
            if not p: return await _json({"ok":False,"error":"Mahsulot topilmadi"})
            if qty > p.stock: return await _json({"ok":False,"error":"Stok yetarli emas"})
            if qty <= 0: return await _json({"ok":False,"error":"Miqdor 0 dan katta bo'lishi kerak"})
            await do_sale(s, p, qty, None, price, note)
            return await _json({"ok":True,"new_stock":p.stock})

    # POST /api/supply
    async def api_supply(req):
        try: data = await req.json()
        except: return await _json({"ok":False,"error":"JSON xato"}, 400)
        barcode   = str(data.get("barcode","")).strip().upper()
        qty       = int(data.get("qty",1))
        buy_price = float(data.get("buy_price",0))
        note      = str(data.get("note","Minisite kirim"))
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
            if not p: return await _json({"ok":False,"error":"Mahsulot topilmadi"})
            if qty <= 0: return await _json({"ok":False,"error":"Miqdor 0 dan katta"})
            await do_supply(s, p, qty, buy_price or p.buy_price, None, note)
            return await _json({"ok":True,"new_stock":p.stock})

    # POST /api/edit
    async def api_edit(req):
        try: data = await req.json()
        except: return await _json({"ok":False,"error":"JSON xato"}, 400)
        barcode = str(data.get("barcode","")).strip().upper()
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
            if not p: return await _json({"ok":False,"error":"Mahsulot topilmadi"})
            for field in ["name","category","color","size","description","unit"]:
                if field in data and data[field]:
                    setattr(p, field, data[field])
            for field in ["buy_price","sell_price","stock","min_stock"]:
                if field in data and data[field] is not None:
                    setattr(p, field, float(data[field]) if "price" in field else int(data[field]))
            p.updated_at = datetime.utcnow()
            await s.commit()
        return await _json({"ok":True})

    # POST /api/archive
    async def api_archive(req):
        try: data = await req.json()
        except: return await _json({"ok":False,"error":"JSON xato"}, 400)
        barcode = str(data.get("barcode","")).strip().upper()
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
            if not p: return await _json({"ok":False,"error":"Topilmadi"})
            p.is_active = False
            await s.commit()
        return await _json({"ok":True})

    # GET /api/history?barcode=
    async def api_history(req):
        barcode = req.rel_url.query.get("barcode","").strip().upper()
        async with AsyncSessionLocal() as s:
            p = await get_product(s, barcode)
            if not p: return await _json({"items":[]})
            # Sales
            sq = await s.execute(
                select(Sale).where(Sale.product_id==p.id)
                .order_by(Sale.sold_at.desc()).limit(20)
            )
            items = []
            for sale in sq.scalars().all():
                items.append({"type":"sell","qty":sale.quantity,"price":sale.sell_price,
                              "total":sale.sell_price*sale.quantity,
                              "date":sale.sold_at.strftime("%d.%m.%Y %H:%M"),"note":sale.note})
            # Supplies
            spq = await s.execute(
                select(Supply).where(Supply.product_id==p.id)
                .order_by(Supply.added_at.desc()).limit(10)
            )
            for sup in spq.scalars().all():
                items.append({"type":"in","qty":sup.quantity,"price":sup.buy_price,
                              "total":sup.buy_price*abs(sup.quantity),
                              "date":sup.added_at.strftime("%d.%m.%Y %H:%M"),"note":sup.note})
            items.sort(key=lambda x: x["date"], reverse=True)
        return await _json({"items":items[:30]})

    # GET /api/stats — umumiy statistika
    async def api_stats(req):
        now = datetime.now()
        async with AsyncSessionLocal() as s:
            r = await s.execute(
                select(func.count(Sale.id), func.sum(Sale.quantity),
                       func.sum(Sale.sell_price*Sale.quantity), func.sum(Sale.profit))
                .where(extract("year",Sale.sold_at)==now.year,
                       extract("month",Sale.sold_at)==now.month)
            )
            row = r.one()
            pc  = await s.execute(select(func.count(Product.id)).where(Product.is_active==True))
            lows= await get_low_stock(s)
        return await _json({
            "month_sales": row[0] or 0,"month_qty": row[1] or 0,
            "month_revenue": row[2] or 0,"month_profit": row[3] or 0,
            "total_products": pc.scalar() or 0,
            "low_stock_count": len(lows),
        })

    # App setup
    from aiohttp import web
    app = web.Application(client_max_size=50*1024*1024)
    app.router.add_get("/health",      health)
    app.router.add_get("/scanner",     scanner)
    app.router.add_get("/p/{barcode}", public_page)
    app.router.add_get("/p",           public_page)
    app.router.add_get("/api/product", api_product)
    app.router.add_get("/api/history", api_history)
    app.router.add_get("/api/stats",   api_stats)
    app.router.add_post("/api/sell",    api_sell)
    app.router.add_post("/api/supply",  api_supply)
    app.router.add_post("/api/edit",    api_edit)
    app.router.add_post("/api/archive", api_archive)

    try:
        cors = aiohttp_cors.setup(app, defaults={"*": aiohttp_cors.ResourceOptions(
            allow_credentials=True, expose_headers="*", allow_headers="*"
        )})
        for route in list(app.router.routes()):
            cors.add(route)
    except Exception as e:
        logger.warning(f"CORS xato: {e}")

    runner = web.AppRunner(app)
    await runner.setup()
    site   = web.TCPSite(runner, "0.0.0.0", WEBAPP_PORT)
    await site.start()
    logger.info(f"🌐 HTTP: http://0.0.0.0:{WEBAPP_PORT}")

# ════════════════════════════════════════════════════════════════
#  6. MIDDLEWARES
# ════════════════════════════════════════════════════════════════
from aiogram import BaseMiddleware
from aiogram.types import Message as AMsg

class DbMiddleware(BaseMiddleware):
    async def __call__(self, handler, event, data):
        async with AsyncSessionLocal() as session:
            data["session"] = session
            return await handler(event, data)

class AuthMiddleware(BaseMiddleware):
    async def __call__(self, handler, event, data):
        user = data.get("event_from_user")
        if not user: return None
        if user.id in ADMIN_IDS: return await handler(event, data)
        session = data.get("session")
        if session:
            worker = await get_worker(session, user.id)
            if worker:
                data["worker"] = worker
                return await handler(event, data)
        # QR deep-link uchun ruxsat (hamma ko'ra oladi)
        msg = event if isinstance(event, AMsg) else None
        if msg and msg.text and msg.text.startswith("/start"):
            args = msg.text.split(maxsplit=1)
            if len(args)>1 and args[1].startswith("qr_"):
                return await handler(event, data)
        if isinstance(event, AMsg):
            await event.answer("⛔️ Siz tizimga qo'shilmagansiz.\n📞 +998909182186")
        return None

class RateLimitMiddleware(BaseMiddleware):
    def __init__(self, limit=30, period=60.0):
        self.limit=limit; self.period=period; self._rec=defaultdict(list)
    async def __call__(self, handler, event, data):
        user=data.get("event_from_user")
        if not user: return await handler(event, data)
        uid=user.id; now=time.monotonic()
        self._rec[uid]=[t for t in self._rec[uid] if now-t<self.period]
        if len(self._rec[uid])>=self.limit:
            if isinstance(event,AMsg): await event.answer("⚠️ Sekinroq!")
            return None
        self._rec[uid].append(now)
        return await handler(event, data)

# ════════════════════════════════════════════════════════════════
#  7. FSM STATES
# ════════════════════════════════════════════════════════════════
from aiogram.fsm.state import State, StatesGroup

class AddSG(StatesGroup):
    photo=State(); code=State(); name=State(); category=State()
    color=State(); size=State(); unit=State()
    buy_price=State(); sell_price=State(); stock=State(); min_stock=State(); confirm=State()

class SaleSG(StatesGroup):
    code=State(); qty=State(); confirm=State()

class SupplySG(StatesGroup):
    code=State(); qty=State(); buy_price=State(); confirm=State()

class EditSG(StatesGroup):
    value=State()

class SearchSG(StatesGroup):
    query=State()

class PostSG(StatesGroup):
    select=State(); caption=State(); confirm=State()

class WorkerSG(StatesGroup):
    tid=State(); name=State(); phone=State(); role=State(); confirm=State()

class WriteoffSG(StatesGroup):
    code=State(); qty=State(); reason=State(); confirm=State()

# ════════════════════════════════════════════════════════════════
#  8. KEYBOARDS
# ════════════════════════════════════════════════════════════════
from aiogram.types import (
    InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, WebAppInfo
)
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder

def main_kb(admin=True) -> ReplyKeyboardMarkup:
    b = ReplyKeyboardBuilder()
    if admin:
        b.row(KeyboardButton(text="📦 Mahsulot qo'shish"), KeyboardButton(text="🔍 Qidirish"))
        b.row(KeyboardButton(text="📤 Sotish"),            KeyboardButton(text="📥 Kirim"))
        b.row(KeyboardButton(text="📢 Kanalga post"),      KeyboardButton(text="📊 Hisobot"))
        b.row(KeyboardButton(text="⚠️ Kam qolganlar"),    KeyboardButton(text="📋 Mahsulotlar"))
        b.row(KeyboardButton(text="👥 Xodimlar"),          KeyboardButton(text="🗑 Spisanie"))
        b.row(KeyboardButton(text="📦 Eksport"),           KeyboardButton(text="🏆 Reyting"))
    else:
        b.row(KeyboardButton(text="🔍 Qidirish"),       KeyboardButton(text="📤 Sotish"))
        b.row(KeyboardButton(text="⚠️ Kam qolganlar"), KeyboardButton(text="📋 Mening sotuvlarim"))
    b.row(KeyboardButton(text="📱 QR Skaner",
          web_app=WebAppInfo(url=f"{WEBAPP_PUBLIC_URL}/scanner")))
    return b.as_markup(resize_keyboard=True)

def cancel_kb():
    b = ReplyKeyboardBuilder()
    b.button(text="❌ Bekor qilish")
    return b.as_markup(resize_keyboard=True)

def skip_kb():
    b = ReplyKeyboardBuilder()
    b.row(KeyboardButton(text="⏭ O'tkazish"), KeyboardButton(text="❌ Bekor qilish"))
    return b.as_markup(resize_keyboard=True)

def confirm_kb():
    b = InlineKeyboardBuilder()
    b.button(text="✅ Tasdiqlash", callback_data="yes")
    b.button(text="❌ Bekor",      callback_data="no")
    b.adjust(2)
    return b.as_markup()

def product_kb(pid: int, admin=True) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    b.button(text="📤 Sotish", callback_data=f"sell:{pid}")
    b.button(text="📱 QR Kod", callback_data=f"qr:{pid}")
    if admin:
        b.button(text="📥 Kirim",      callback_data=f"supply:{pid}")
        b.button(text="✏️ Tahrirlash", callback_data=f"edit:{pid}")
        b.button(text="📢 Kanalga",    callback_data=f"post:{pid}")
        b.button(text="🖨 Label 58mm", callback_data=f"label:{pid}")
        b.button(text="📊 Tarix",      callback_data=f"hist:{pid}")
        b.button(text="🗑 O'chirish",  callback_data=f"del:{pid}")
        b.button(text="🌐 Public link",callback_data=f"link:{pid}")
        b.adjust(2,2,2,2,1)
    else:
        b.adjust(2)
    return b.as_markup()

def edit_kb(pid: int) -> InlineKeyboardMarkup:
    b = InlineKeyboardBuilder()
    for label, field in [
        ("📝 Nomi","name"), ("💵 Sotish narxi","sell_price"),
        ("💰 Tannarx","buy_price"), ("📦 Stok","stock"),
        ("⚠️ Min stok","min_stock"), ("🎨 Rang","color"),
        ("📐 O'lcham","size"), ("📄 Tavsif","description"),
    ]:
        b.button(text=label, callback_data=f"ef:{pid}:{field}")
    b.button(text="⬅️ Orqaga", callback_data=f"back:{pid}")
    b.adjust(2)
    return b.as_markup()

def report_kb() -> InlineKeyboardMarkup:
    now=datetime.now()
    py,pm=(now.year-1,12) if now.month==1 else (now.year,now.month-1)
    b=InlineKeyboardBuilder()
    b.button(text=f"📅 Bu oy",       callback_data=f"rpt:{now.year}:{now.month}")
    b.button(text="📅 O'tgan oy",    callback_data=f"rpt:{py}:{pm}")
    b.button(text="📊 Bugun",        callback_data="rpt_today")
    b.button(text="📥 Excel",        callback_data=f"rpt_xl:{now.year}:{now.month}")
    b.adjust(2,2)
    return b.as_markup()

# ════════════════════════════════════════════════════════════════
#  9. ROUTER & HANDLERS
# ════════════════════════════════════════════════════════════════
from aiogram import Router, F, Bot
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, Message, BufferedInputFile

router = Router()
for mw in [DbMiddleware(), RateLimitMiddleware(), AuthMiddleware()]:
    router.message.middleware(mw)
    router.callback_query.middleware(mw)

async def _adm(msg, s): return await is_admin(s, msg.from_user.id)

async def send_product(msg, p: Product, admin=True):
    kb  = product_kb(p.id, admin=admin)
    txt = product_text(p, admin)
    if p.photo_file_id:
        await msg.answer_photo(p.photo_file_id, caption=txt, parse_mode="HTML", reply_markup=kb)
    else:
        await msg.answer(txt, parse_mode="HTML", reply_markup=kb)

async def send_public_product(msg, p: Product):
    """Xaridor uchun — stok ko'rsatilmaydi"""
    b = InlineKeyboardBuilder()
    b.button(text="🛍 Do'konga o'tish",  url=f"https://t.me/{CHANNEL_ID.lstrip('@')}")
    b.button(text="🌐 Mahsulot sahifasi", url=public_url(p))
    b.adjust(1)
    lines = [f"🛍 <b>{p.name}</b>", ""]
    if p.category: lines.append(f"🗂 {p.category}")
    if p.color:    lines.append(f"🎨 Rang: {p.color}")
    if p.size:     lines.append(f"📐 O'lcham: {p.size}")
    lines += ["", f"💵 Narx: <b>{fmt(p.sell_price)} so'm</b>", ""]
    if p.description: lines += [f"📄 {p.description}", ""]
    lines += ["📞 +998909182186", "📞 +998947043111", f"🌐 {CHANNEL_ID}"]
    caption = "\n".join(lines)
    if p.photo_file_id:
        await msg.answer_photo(p.photo_file_id, caption=caption, parse_mode="HTML", reply_markup=b.as_markup())
    else:
        await msg.answer(caption, parse_mode="HTML", reply_markup=b.as_markup())

async def auto_send_qr(msg, bot, p: Product):
    """Mahsulot qo'shilganda avtomatik QR + label yuborish"""
    if not p.barcode: return
    qr = make_product_qr(p)
    if qr:
        await msg.answer_photo(
            BufferedInputFile(qr, f"qr_{p.code}.png"),
            caption=(
                f"📱 <b>QR Kod tayyor!</b>\n\n"
                f"📦 {p.name}\n"
                f"📊 Barcode: <code>{p.barcode}</code>\n"
                f"🌐 <a href='{public_url(p)}'>{public_url(p)}</a>\n\n"
                f"⬆️ Chop etib mahsulotga yopishtirib qo'ying."
            ),
            parse_mode="HTML"
        )
    label = make_label_58mm(p)
    if label:
        await msg.answer_photo(
            BufferedInputFile(label, f"label_{p.code}.png"),
            caption=f"🏷 <b>58mm Label</b> — termal printer uchun",
            parse_mode="HTML"
        )

# ── /start ───────────────────────────────────────────────────────────────────
@router.message(CommandStart())
async def cmd_start(msg: Message, session: AsyncSession, state: FSMContext):
    await state.clear()
    adm    = await _adm(msg, session)
    worker = await get_worker(session, msg.from_user.id)

    # Deep-link: /start qr_SAR-123456
    args = msg.text.split(maxsplit=1)
    if len(args)>1 and args[1].startswith("qr_"):
        barcode = args[1][3:]
        p = await get_product(session, barcode)
        if p:
            if adm or worker: await send_product(msg, p, admin=adm)
            else: await send_public_product(msg, p)
        else:
            await msg.answer(
                f"❌ <b>Mahsulot topilmadi</b>\n📊 Barcode: <code>{barcode}</code>",
                parse_mode="HTML"
            )
        return

    if not adm and not worker:
        await msg.answer(
            f"👋 Salom, <b>{msg.from_user.first_name}</b>!\n\n"
            f"🏪 <b>Sariosiyo Online</b> do'konimizga xush kelibsiz!\n\n"
            f"📱 Mahsulot QR kodini skanerlash uchun:\n"
            f"🌐 {WEBAPP_PUBLIC_URL}\n\n"
            f"📞 +998909182186\n"
            f"🌐 {CHANNEL_ID}",
            parse_mode="HTML"
        )
        return

    kb   = main_kb(admin=adm)
    role = "Admin" if adm else f"Xodim: {worker.name}"
    await msg.answer(
        f"👋 Salom, <b>{msg.from_user.first_name}</b>! [{role}]\n"
        f"🏪 <b>Qaytganingiz uchun Rahmat</b>\n\n"
        f"🏪 <b>Bugun nimalar qilamiz Nima vazifalaringiz bor</b>\n\n"
        f"🌐 Server: {WEBAPP_PUBLIC_URL}",
        reply_markup=kb, parse_mode="HTML"
    )

@router.message(F.text == "❌ Bekor qilish")
async def cancel(msg: Message, state: FSMContext, session: AsyncSession):
    await state.clear()
    adm = await _adm(msg, session)
    await msg.answer("❌ Bekor qilindi.", reply_markup=main_kb(admin=adm))

@router.message(F.web_app_data)
async def webapp_data(msg: Message, session: AsyncSession, state: FSMContext):
    try:
        data = json.loads(msg.web_app_data.data)
        code = data.get("code") or data.get("barcode","")
    except:
        await msg.answer("⚠️ Noto'g'ri ma'lumot.")
        return
    if not code: return
    adm = await _adm(msg, session)
    p   = await get_product(session, code.upper())
    if not p:
        await msg.answer(f"❌ Topilmadi: <code>{code.upper()}</code>", parse_mode="HTML")
        return
    if data.get("action") == "sell":
        await state.set_state(SaleSG.qty)
        await state.update_data(product_id=p.id)
        await msg.answer(
            f"📤 <b>{p.name}</b>\n💵 {fmt(p.sell_price)} so'm\n"
            f"📦 Stok: <b>{p.stock} {p.unit}</b>\n\nNechta?",
            reply_markup=cancel_kb(), parse_mode="HTML"
        )
    else:
        await send_product(msg, p, admin=adm)

# ── MAHSULOT QO'SHISH ────────────────────────────────────────────────────────
@router.message(F.text == "📦 Mahsulot qo'shish")
async def add_start(msg: Message, state: FSMContext, session: AsyncSession):
    if not await _adm(msg, session): return
    await state.set_state(AddSG.photo)
    await msg.answer(
        f"📦 <b>Yangi Mahsulot</b>\n{_progress(1)}\n\n"
        f"📸 <b>Rasm</b> yuboring (yoki o'tkazib yuboring):",
        reply_markup=skip_kb(), parse_mode="HTML"
    )

@router.message(AddSG.photo, F.photo)
async def add_photo(msg: Message, state: FSMContext):
    await state.update_data(photo_file_id=msg.photo[-1].file_id)
    await state.set_state(AddSG.code)
    await msg.answer(
        f"✅ Rasm saqlandi!\n{_progress(2)}\n\n"
        f"🔖 <b>Ichki kod</b> kiriting:\n<i>Misol: DB-98-44, KYM-001</i>",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )

@router.message(AddSG.photo, F.text == "⏭ O'tkazish")
async def add_photo_skip(msg: Message, state: FSMContext):
    await state.update_data(photo_file_id="")
    await state.set_state(AddSG.code)
    await msg.answer(f"{_progress(2)}\n\n🔖 <b>Ichki kod</b>:\n<i>Misol: DB-98-44</i>",
                     reply_markup=cancel_kb(), parse_mode="HTML")

@router.message(AddSG.photo)
async def add_photo_err(msg: Message):
    await msg.answer("⚠️ Rasm yuboring yoki «⏭ O'tkazish» bosing.")

@router.message(AddSG.code)
async def add_code(msg: Message, state: FSMContext, session: AsyncSession):
    code = msg.text.strip().upper()
    if await get_product(session, code):
        await msg.answer(f"⚠️ <code>{code}</code> allaqachon mavjud!"); return
    barcode = await unique_barcode(session)
    await state.update_data(code=code, barcode=barcode)
    await state.set_state(AddSG.name)
    await msg.answer(
        f"✅ Kod: <code>{code}</code>\n"
        f"🎲 Barcode (avto): <code>{barcode}</code>\n\n"
        f"{_progress(3)}\n\n📝 <b>Mahsulot nomi</b>:",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )

@router.message(AddSG.name)
async def add_name(msg: Message, state: FSMContext):
    await state.update_data(name=msg.text.strip())
    await state.set_state(AddSG.category)
    b = ReplyKeyboardBuilder()
    for c in ["Kiyim","Poyabzal","Sumka","Aksessuar","Sport","Kosmetika","Elektronika","Boshqa"]:
        b.button(text=c)
    b.button(text="⏭ O'tkazish"); b.button(text="❌ Bekor qilish"); b.adjust(4)
    await msg.answer(f"{_progress(4)}\n\n🗂 <b>Kategoriya</b>:",
                     reply_markup=b.as_markup(resize_keyboard=True), parse_mode="HTML")

@router.message(AddSG.category)
async def add_category(msg: Message, state: FSMContext):
    await state.update_data(category="" if msg.text=="⏭ O'tkazish" else msg.text.strip())
    await state.set_state(AddSG.color)
    b = ReplyKeyboardBuilder()
    for c in ["Qora","Oq","Qizil","Ko'k","Yashil","Sariq","Jigarrang","Kulrang","Aralash"]:
        b.button(text=c)
    b.button(text="⏭ O'tkazish"); b.button(text="❌ Bekor qilish"); b.adjust(3)
    await msg.answer(f"{_progress(5)}\n\n🎨 <b>Rang</b>:",
                     reply_markup=b.as_markup(resize_keyboard=True), parse_mode="HTML")

@router.message(AddSG.color)
async def add_color(msg: Message, state: FSMContext):
    await state.update_data(color="" if msg.text=="⏭ O'tkazish" else msg.text.strip())
    await state.set_state(AddSG.size)
    b = ReplyKeyboardBuilder()
    for s in ["XS","S","M","L","XL","XXL","XXXL"]:
        b.button(text=s)
    for s in ["36","37","38","39","40","41","42","43","44"]:
        b.button(text=s)
    b.button(text="⏭ O'tkazish"); b.button(text="❌ Bekor qilish"); b.adjust(7,9)
    await msg.answer(f"{_progress(6)}\n\n📐 <b>O'lcham</b>:",
                     reply_markup=b.as_markup(resize_keyboard=True), parse_mode="HTML")

@router.message(AddSG.size)
async def add_size(msg: Message, state: FSMContext):
    await state.update_data(size="" if msg.text=="⏭ O'tkazish" else msg.text.strip())
    await state.set_state(AddSG.unit)
    b = ReplyKeyboardBuilder()
    for u in ["dona","juft","xil","to'plam","metr","kg","litr"]:
        b.button(text=u)
    b.adjust(4)
    await msg.answer(f"{_progress(7)}\n\n📏 <b>Birlik</b>:",
                     reply_markup=b.as_markup(resize_keyboard=True), parse_mode="HTML")

@router.message(AddSG.unit)
async def add_unit(msg: Message, state: FSMContext):
    await state.update_data(unit=msg.text.strip())
    await state.set_state(AddSG.buy_price)
    await msg.answer(f"{_progress(8)}\n\n💰 <b>Tannarx</b> (so'm):",
                     reply_markup=cancel_kb(), parse_mode="HTML")

@router.message(AddSG.buy_price)
async def add_buy(msg: Message, state: FSMContext):
    try:
        p = float(msg.text.strip().replace(" ","").replace(",","")); assert p>=0
    except:
        await msg.answer("⚠️ Raqam kiriting"); return
    await state.update_data(buy_price=p)
    await state.set_state(AddSG.sell_price)
    await msg.answer(f"✅ Tannarx: <b>{fmt(p)} so'm</b>\n\n"
                     f"{_progress(9)}\n\n💵 <b>Sotish narxi</b>:", parse_mode="HTML")

@router.message(AddSG.sell_price)
async def add_sell(msg: Message, state: FSMContext):
    try:
        p = float(msg.text.strip().replace(" ","").replace(",","")); assert p>=0
    except:
        await msg.answer("⚠️ Raqam kiriting"); return
    d  = await state.get_data()
    bp = d.get("buy_price",0)
    pf = p - bp; pct = (pf/bp*100) if bp>0 else 0
    await state.update_data(sell_price=p)
    await state.set_state(AddSG.stock)
    await msg.answer(
        f"✅ Sotish: <b>{fmt(p)} so'm</b>\n"
        f"📈 Foyda: <b>{fmt(pf)} so'm ({pct:.0f}%)</b>\n\n"
        f"{_progress(10)}\n\n📦 <b>Boshlang'ich stok</b>:", parse_mode="HTML"
    )

@router.message(AddSG.stock)
async def add_stock(msg: Message, state: FSMContext):
    try:
        s = int(msg.text.strip()); assert s>=0
    except:
        await msg.answer("⚠️ Butun son"); return
    await state.update_data(stock=s)
    await state.set_state(AddSG.min_stock)
    await msg.answer(f"{_progress(11)}\n\n⚠️ <b>Minimal qoldiq</b> (ogohlantirish chegarasi):\n"
                     f"<i>Default: {LOW_STOCK_ALERT} ta</i>",
                     reply_markup=skip_kb(), parse_mode="HTML")

@router.message(AddSG.min_stock)
async def add_min_stock(msg: Message, state: FSMContext):
    ms = LOW_STOCK_ALERT if msg.text=="⏭ O'tkazish" else int(msg.text.strip()) if msg.text.strip().isdigit() else LOW_STOCK_ALERT
    await state.update_data(min_stock=ms)
    await state.set_state(AddSG.confirm)
    d  = await state.get_data()
    pf = d['sell_price']-d['buy_price']; pct=(pf/d['buy_price']*100) if d['buy_price']>0 else 0
    await msg.answer(
        f"📋 <b>Tasdiqlang:</b>\n\n"
        f"{'📸 Rasm: ✅' if d.get('photo_file_id') else '📸 Rasm: —'}\n"
        f"🔖 Kod: <code>{d['code']}</code>\n"
        f"📊 Barcode: <code>{d.get('barcode','—')}</code>\n"
        f"📝 Nomi: <b>{d['name']}</b>\n"
        f"🗂 {d.get('category') or '—'}\n"
        f"🎨 {d.get('color') or '—'}\n"
        f"📐 {d.get('size') or '—'}\n"
        f"📏 {d['unit']}\n"
        f"💰 Tannarx: {fmt(d['buy_price'])} so'm\n"
        f"💵 Sotish: <b>{fmt(d['sell_price'])} so'm</b>\n"
        f"📈 Foyda: {fmt(pf)} ({pct:.0f}%)\n"
        f"📦 Stok: {d['stock']} {d['unit']}\n"
        f"⚠️ Min: {ms}",
        reply_markup=confirm_kb(), parse_mode="HTML"
    )

@router.callback_query(AddSG.confirm, F.data == "yes")
async def add_confirm(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    d = await state.get_data()
    p = Product(
        code=d["code"], barcode=d.get("barcode",""), name=d["name"],
        category=d.get("category",""), color=d.get("color",""), size=d.get("size",""),
        unit=d["unit"], buy_price=d["buy_price"], sell_price=d["sell_price"],
        stock=d["stock"], min_stock=d["min_stock"], photo_file_id=d.get("photo_file_id","")
    )
    session.add(p)
    await session.flush()
    if d["stock"]>0:
        session.add(Supply(product_id=p.id, quantity=d["stock"],
                           buy_price=d["buy_price"], note="Boshlang'ich kirim"))
    await session.commit()
    await state.clear()

    await cb.message.answer(
        f"✅ <b>Mahsulot qo'shildi!</b>\n\n" + product_text(p),
        parse_mode="HTML", reply_markup=product_kb(p.id)
    )
    await auto_send_qr(cb.message, cb.bot, p)
    await cb.message.answer("📋 Menyu:", reply_markup=main_kb())
    await cb.answer()
    await check_low(cb.bot, p)

# ── QIDIRUV ──────────────────────────────────────────────────────────────────
@router.message(F.text == "🔍 Qidirish")
async def search_start(msg: Message, state: FSMContext):
    await state.set_state(SearchSG.query)
    await msg.answer("🔍 Qidirish — kod, barcode, nom:", reply_markup=cancel_kb())

@router.message(SearchSG.query)
async def search_do(msg: Message, state: FSMContext, session: AsyncSession):
    q   = msg.text.strip()
    adm = await _adm(msg, session)
    await state.clear()
    p   = await get_product(session, q)
    if p:
        await send_product(msg, p, admin=adm)
        return
    prods = await search_products(session, q)
    if not prods:
        await msg.answer(f"😕 <b>«{q}»</b> bo'yicha topilmadi.",
                         reply_markup=main_kb(admin=adm), parse_mode="HTML")
        return
    if len(prods)==1:
        await send_product(msg, prods[0], admin=adm)
        return
    b   = InlineKeyboardBuilder()
    txt = f"🔍 <b>{len(prods)} ta natija:</b>\n\n"
    for pr in prods:
        icon = "🔴" if pr.stock<=pr.min_stock else "🟢"
        txt += f"{icon} <code>{pr.code}</code> — {pr.name} ({pr.stock} {pr.unit})\n"
        b.button(text=f"{pr.code}|{pr.name[:22]}", callback_data=f"sp:{pr.id}")
    b.adjust(1)
    await msg.answer(txt, parse_mode="HTML", reply_markup=b.as_markup())

@router.callback_query(F.data.startswith("sp:"))
async def show_cb(cb: CallbackQuery, session: AsyncSession):
    adm = await is_admin(session, cb.from_user.id)
    p   = await session.get(Product, int(cb.data.split(":")[1]))
    if p: await send_product(cb.message, p, admin=adm)
    await cb.answer()

# ── SOTISH ───────────────────────────────────────────────────────────────────
@router.message(F.text == "📤 Sotish")
async def sale_start(msg: Message, state: FSMContext):
    await state.set_state(SaleSG.code)
    await msg.answer("📤 Mahsulot kodi:", reply_markup=cancel_kb())

@router.callback_query(F.data.startswith("sell:"))
async def sell_cb(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p: await cb.answer("Topilmadi."); return
    if p.stock<=0: await cb.answer(f"❌ {p.name} tugagan!", show_alert=True); return
    await state.set_state(SaleSG.qty)
    await state.update_data(product_id=p.id)
    await cb.message.answer(
        f"📤 <b>{p.name}</b>\n💵 {fmt(p.sell_price)} so'm\n"
        f"📦 Stok: <b>{p.stock}</b>\n\nNechta?",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )
    await cb.answer()

@router.message(SaleSG.code)
async def sale_code(msg: Message, state: FSMContext, session: AsyncSession):
    p = await get_product(session, msg.text)
    if not p: await msg.answer(f"⚠️ Topilmadi: <code>{msg.text.upper()}</code>"); return
    if p.stock<=0: await msg.answer(f"❌ <b>{p.name}</b> — stokda yo'q!"); return
    await state.update_data(product_id=p.id)
    await state.set_state(SaleSG.qty)
    await msg.answer(f"✅ <b>{p.name}</b>\n💵 {fmt(p.sell_price)}\n📦 Stok: <b>{p.stock}</b>\n\nNechta?",
                     parse_mode="HTML")

@router.message(SaleSG.qty)
async def sale_qty(msg: Message, state: FSMContext, session: AsyncSession):
    try:
        qty=int(msg.text.strip()); assert qty>0
    except:
        await msg.answer("⚠️ Musbat butun son!"); return
    d = await state.get_data()
    p = await session.get(Product, d["product_id"])
    if qty>p.stock:
        await msg.answer(f"❌ Faqat <b>{p.stock}</b> {p.unit} bor!", parse_mode="HTML"); return
    await state.update_data(qty=qty)
    await state.set_state(SaleSG.confirm)
    total = p.sell_price*qty
    await msg.answer(
        f"📤 <b>Tasdiqlash:</b>\n\n"
        f"📦 {p.name}\n"
        f"🔢 {fmt(p.sell_price)} × {qty} = <b>{fmt(total)} so'm</b>\n"
        f"📦 Qoladi: {p.stock-qty} {p.unit}",
        reply_markup=confirm_kb(), parse_mode="HTML"
    )

@router.callback_query(SaleSG.confirm, F.data == "yes")
async def sale_confirm(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    d      = await state.get_data()
    p      = await session.get(Product, d["product_id"])
    worker = await get_worker(session, cb.from_user.id)
    sale   = await do_sale(session, p, d["qty"], worker)
    await state.clear()
    total  = sale.sell_price*sale.quantity
    await cb.message.answer(
        f"✅ <b>Sotildi!</b>\n\n"
        f"📦 {p.name} × {sale.quantity} {p.unit}\n"
        f"💵 Jami: <b>{fmt(total)} so'm</b>\n"
        f"📦 Qoldi: <b>{p.stock} {p.unit}</b>",
        reply_markup=main_kb(admin=await is_admin(session, cb.from_user.id)),
        parse_mode="HTML"
    )
    await cb.answer("✅")
    await check_low(cb.bot, p)

# ── KIRIM ────────────────────────────────────────────────────────────────────
@router.message(F.text == "📥 Kirim")
async def supply_start(msg: Message, state: FSMContext, session: AsyncSession):
    if not await _adm(msg, session): return
    await state.set_state(SupplySG.code)
    await msg.answer("📥 Mahsulot kodi:", reply_markup=cancel_kb())

@router.callback_query(F.data.startswith("supply:"))
async def supply_cb(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    if not await is_admin(session, cb.from_user.id): await cb.answer("⛔️"); return
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p: await cb.answer("Topilmadi."); return
    await state.set_state(SupplySG.qty)
    await state.update_data(product_id=p.id, buy_price=p.buy_price)
    await cb.message.answer(
        f"📥 <b>{p.name}</b>\nStok: {p.stock}\n\nNechta keldi?",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )
    await cb.answer()

@router.message(SupplySG.code)
async def supply_code(msg: Message, state: FSMContext, session: AsyncSession):
    p = await get_product(session, msg.text)
    if not p: await msg.answer("⚠️ Topilmadi."); return
    await state.update_data(product_id=p.id, buy_price=p.buy_price)
    await state.set_state(SupplySG.qty)
    await msg.answer(f"✅ <b>{p.name}</b>\nStok: {p.stock}\n\nNechta?", parse_mode="HTML")

@router.message(SupplySG.qty)
async def supply_qty(msg: Message, state: FSMContext):
    try:
        qty=int(msg.text.strip()); assert qty>0
    except:
        await msg.answer("⚠️ Musbat son!"); return
    await state.update_data(qty=qty)
    await state.set_state(SupplySG.buy_price)
    d = await state.get_data()
    await msg.answer(f"💰 Tannarx?\nOldingi: <b>{fmt(d['buy_price'])} so'm</b>\n"
                     f"O'zgarmasa «O'sha» yozing:", parse_mode="HTML")

@router.message(SupplySG.buy_price)
async def supply_bp(msg: Message, state: FSMContext, session: AsyncSession):
    d = await state.get_data()
    if msg.text.strip().lower() in ["o'sha","same","-","="]:
        bp = d["buy_price"]
    else:
        try:
            bp=float(msg.text.strip().replace(" ","").replace(",","")); assert bp>=0
        except:
            await msg.answer("⚠️ Raqam!"); return
    p = await session.get(Product, d["product_id"])
    await state.update_data(buy_price=bp)
    await state.set_state(SupplySG.confirm)
    await msg.answer(
        f"📥 <b>Tasdiqlash:</b>\n\n"
        f"📦 {p.name} × {d['qty']} {p.unit}\n"
        f"💰 {fmt(bp)} so'm/dona\n"
        f"💸 Jami: <b>{fmt(bp*d['qty'])} so'm</b>\n"
        f"Yangi stok: <b>{p.stock+d['qty']}</b>",
        reply_markup=confirm_kb(), parse_mode="HTML"
    )

@router.callback_query(SupplySG.confirm, F.data == "yes")
async def supply_confirm(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    d = await state.get_data()
    p = await session.get(Product, d["product_id"])
    w = await get_worker(session, cb.from_user.id)
    await do_supply(session, p, d["qty"], d["buy_price"], w)
    await state.clear()
    await cb.message.answer(
        f"✅ <b>Kirim!</b>\n📦 {p.name} +{d['qty']} {p.unit}\n"
        f"Yangi stok: <b>{p.stock}</b>",
        reply_markup=main_kb(), parse_mode="HTML"
    )
    await cb.answer("✅")

# ── QR ───────────────────────────────────────────────────────────────────────
@router.callback_query(F.data.startswith("qr:"))
async def qr_cb(cb: CallbackQuery, session: AsyncSession):
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p or not p.barcode: await cb.answer("Barcode yo'q!"); return
    await cb.answer("⏳ QR yasalmoqda...")
    qr = make_product_qr(p)
    if qr:
        await cb.message.answer_photo(
            BufferedInputFile(qr, f"qr_{p.code}.png"),
            caption=f"📱 <b>QR Kod</b>\n\n📦 {p.name}\n"
                    f"📊 <code>{p.barcode}</code>\n"
                    f"🌐 <a href='{public_url(p)}'>{public_url(p)}</a>",
            parse_mode="HTML"
        )
    else:
        await cb.message.answer("⚠️ pip install qrcode Pillow")

@router.callback_query(F.data.startswith("label:"))
async def label_cb(cb: CallbackQuery, session: AsyncSession):
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p: await cb.answer("Topilmadi."); return
    await cb.answer("⏳")
    lb = make_label_58mm(p)
    if lb:
        await cb.message.answer_photo(
            BufferedInputFile(lb, f"label_{p.code}.png"),
            caption=f"🏷 <b>58mm Label</b>\n{p.name}", parse_mode="HTML"
        )
    else:
        await cb.message.answer("⚠️ pip install qrcode Pillow")

@router.callback_query(F.data.startswith("link:"))
async def link_cb(cb: CallbackQuery, session: AsyncSession):
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p: await cb.answer("Topilmadi."); return
    url = public_url(p)
    await cb.message.answer(
        f"🌐 <b>Public link:</b>\n{url}\n\n"
        f"📊 Barcode: <code>{p.barcode or p.code}</code>",
        parse_mode="HTML"
    )
    await cb.answer()

# ── TAHRIRLASH ────────────────────────────────────────────────────────────────
@router.callback_query(F.data.startswith("edit:"))
async def edit_start(cb: CallbackQuery, session: AsyncSession):
    if not await is_admin(session, cb.from_user.id): await cb.answer("⛔️"); return
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p: await cb.answer("Topilmadi."); return
    await cb.message.answer(f"✏️ <b>{p.name}</b>:", reply_markup=edit_kb(p.id), parse_mode="HTML")
    await cb.answer()

@router.callback_query(F.data.startswith("ef:"))
async def edit_field(cb: CallbackQuery, state: FSMContext):
    _, pid, field = cb.data.split(":")
    await state.set_state(EditSG.value)
    await state.update_data(product_id=int(pid), field=field)
    labels={"name":"nomi","sell_price":"sotish narxi","buy_price":"tannarx",
             "stock":"stok","min_stock":"min stok","color":"rangi","size":"o'lchami","description":"tavsifi"}
    await cb.message.answer(f"✏️ Yangi <b>{labels.get(field,field)}</b>:",
                             reply_markup=cancel_kb(), parse_mode="HTML")
    await cb.answer()

@router.message(EditSG.value)
async def edit_save(msg: Message, state: FSMContext, session: AsyncSession):
    d = await state.get_data()
    p = await session.get(Product, d["product_id"])
    try:
        if d["field"] in ("sell_price","buy_price"):
            setattr(p, d["field"], float(msg.text.strip().replace(" ","").replace(",","")))
        elif d["field"] in ("stock","min_stock"):
            setattr(p, d["field"], int(msg.text.strip()))
        else:
            setattr(p, d["field"], msg.text.strip())
        p.updated_at = datetime.utcnow()
        await session.commit()
        await state.clear()
        await msg.answer("✅ Yangilandi!", reply_markup=main_kb())
        await send_product(msg, p, admin=True)
    except Exception as e:
        await msg.answer(f"⚠️ Xato: {e}")

# ── POST ─────────────────────────────────────────────────────────────────────
@router.message(F.text == "📢 Kanalga post")
async def post_start(msg: Message, state: FSMContext, session: AsyncSession):
    if not await _adm(msg, session): return
    await state.set_state(PostSG.select)
    await msg.answer("📢 Mahsulot kodi:", reply_markup=cancel_kb())

@router.callback_query(F.data.startswith("post:"))
async def post_cb(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    p = await session.get(Product, int(cb.data.split(":")[1]))
    if not p: await cb.answer("Topilmadi."); return
    await _post_preview(cb.message, state, p)
    await cb.answer()

@router.message(PostSG.select)
async def post_select(msg: Message, state: FSMContext, session: AsyncSession):
    p = await get_product(session, msg.text)
    if not p:
        prods = await search_products(session, msg.text)
        if not prods: await msg.answer("😕 Topilmadi."); return
        p = prods[0] if len(prods)==1 else None
        if not p:
            b = InlineKeyboardBuilder()
            for pr in prods[:8]: b.button(text=f"{pr.code}|{pr.name[:22]}", callback_data=f"post:{pr.id}")
            b.adjust(1)
            await msg.answer("Tanlang:", reply_markup=b.as_markup()); return
    await _post_preview(msg, state, p)

async def _post_preview(msg, state, p):
    cap = channel_caption(p)
    await state.update_data(product_id=p.id, caption=cap)
    await state.set_state(PostSG.confirm)
    b = InlineKeyboardBuilder()
    b.button(text="📢 Yuborish",          callback_data="post_send")
    b.button(text="✏️ Caption o'zgartir", callback_data="post_edit")
    b.button(text="❌ Bekor",             callback_data="post_cancel")
    b.adjust(2,1)
    if p.photo_file_id:
        await msg.answer_photo(p.photo_file_id, caption=f"👁 Preview:\n\n{cap}",
                               parse_mode="HTML", reply_markup=b.as_markup())
    else:
        await msg.answer(f"👁 Preview:\n\n{cap}", parse_mode="HTML", reply_markup=b.as_markup())

@router.callback_query(PostSG.confirm, F.data == "post_send")
async def post_send(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    d = await state.get_data()
    p = await session.get(Product, d["product_id"])
    try:
        if p.photo_file_id:
            sent = await cb.bot.send_photo(CHANNEL_ID, p.photo_file_id, caption=d["caption"], parse_mode="HTML")
        else:
            sent = await cb.bot.send_message(CHANNEL_ID, d["caption"], parse_mode="HTML")
        session.add(ChannelPost(product_id=p.id, message_id=sent.message_id,
                                caption=d["caption"], posted_by=cb.from_user.id))
        await session.commit()
        await state.clear()
        await cb.message.answer(f"✅ Post yuborildi → {CHANNEL_ID}", reply_markup=main_kb())
    except Exception as e:
        await cb.message.answer(f"❌ Xato: {e}")
    await cb.answer()

@router.callback_query(PostSG.confirm, F.data == "post_edit")
async def post_edit(cb: CallbackQuery, state: FSMContext):
    await state.set_state(PostSG.caption)
    await cb.message.answer("✏️ Yangi caption:", reply_markup=cancel_kb())
    await cb.answer()

@router.message(PostSG.caption)
async def post_caption(msg: Message, state: FSMContext, session: AsyncSession):
    await state.update_data(caption=msg.text)
    await state.set_state(PostSG.confirm)
    d = await state.get_data()
    p = await session.get(Product, d["product_id"])
    b = InlineKeyboardBuilder()
    b.button(text="📢 Yuborish", callback_data="post_send")
    b.button(text="❌ Bekor",    callback_data="post_cancel")
    b.adjust(2)
    if p.photo_file_id:
        await msg.answer_photo(p.photo_file_id, caption=msg.text, parse_mode="HTML", reply_markup=b.as_markup())
    else:
        await msg.answer(msg.text, parse_mode="HTML", reply_markup=b.as_markup())

@router.callback_query(PostSG.confirm, F.data == "post_cancel")
async def post_cancel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.answer("❌ Bekor.", reply_markup=main_kb())
    await cb.answer()

# ── HISOBOT ──────────────────────────────────────────────────────────────────
@router.message(F.text == "📊 Hisobot")
async def report(msg: Message, session: AsyncSession):
    if not await _adm(msg, session): return
    await msg.answer("📊 <b>Hisobot</b>", reply_markup=report_kb(), parse_mode="HTML")

@router.callback_query(F.data.startswith("rpt:"))
async def monthly_rpt(cb: CallbackQuery, session: AsyncSession):
    _, y, m = cb.data.split(":"); y,m=int(y),int(m)
    r  = await session.execute(
        select(func.count(Sale.id), func.sum(Sale.quantity),
               func.sum(Sale.sell_price*Sale.quantity), func.sum(Sale.profit))
        .where(extract("year",Sale.sold_at)==y, extract("month",Sale.sold_at)==m)
    )
    row = r.one()
    cnt,qty,rev,prf = row[0] or 0, row[1] or 0, row[2] or 0.0, row[3] or 0.0
    top = await session.execute(
        select(Product.name, func.sum(Sale.quantity).label("q"))
        .join(Sale, Sale.product_id==Product.id)
        .where(extract("year",Sale.sold_at)==y, extract("month",Sale.sold_at)==m)
        .group_by(Product.id).order_by(func.sum(Sale.quantity).desc()).limit(5)
    )
    txt = (f"📊 <b>{calendar.month_name[m]} {y}</b>\n\n"
           f"📦 Sotuvlar: {cnt}\n🔢 Dona: {qty}\n"
           f"💵 Daromad: <b>{fmt(rev)} so'm</b>\n"
           f"📈 Foyda: <b>{fmt(prf)} so'm</b>\n")
    for i,row in enumerate(top.all(),1):
        if i==1: txt+="\n🏆 <b>Top:</b>\n"
        txt+=f"{i}. {row.name[:25]} — {row.q} dona\n"
    await cb.message.edit_text(txt, parse_mode="HTML", reply_markup=report_kb())
    await cb.answer()

@router.callback_query(F.data == "rpt_today")
async def today_rpt(cb: CallbackQuery, session: AsyncSession):
    today=date.today()
    r=await session.execute(
        select(func.count(Sale.id),func.sum(Sale.quantity),
               func.sum(Sale.sell_price*Sale.quantity),func.sum(Sale.profit))
        .where(func.date(Sale.sold_at)==today)
    )
    row=r.one()
    await cb.message.edit_text(
        f"📊 <b>Bugun — {today:%d.%m.%Y}</b>\n\n"
        f"📦 Sotuvlar: {row[0] or 0}\n🔢 Dona: {row[1] or 0}\n"
        f"💵 Daromad: <b>{fmt(row[2] or 0)} so'm</b>\n"
        f"📈 Foyda: <b>{fmt(row[3] or 0)} so'm</b>",
        reply_markup=report_kb(), parse_mode="HTML"
    )
    await cb.answer()

@router.callback_query(F.data.startswith("rpt_xl:"))
async def excel_rpt(cb: CallbackQuery, session: AsyncSession):
    _,y,m=cb.data.split(":"); y,m=int(y),int(m)
    await cb.answer("⏳ Excel...")
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment
        from openpyxl.utils import get_column_letter
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Sotuvlar"
        RED=PatternFill("solid",fgColor="DC1E32")
        hdrs=["Sana","Kod","Nomi","Miqdor","Narx","Jami","Foyda"]
        for c,h in enumerate(hdrs,1):
            cell=ws.cell(1,c,h)
            cell.font=Font(bold=True,color="FFFFFF",size=10)
            cell.fill=RED; cell.alignment=Alignment(horizontal="center")
        q=await session.execute(
            select(Sale,Product.code,Product.name)
            .join(Product,Sale.product_id==Product.id)
            .where(extract("year",Sale.sold_at)==y,extract("month",Sale.sold_at)==m)
            .order_by(Sale.sold_at.desc())
        )
        for ri,(s,code,name) in enumerate(q.all(),2):
            ws.cell(ri,1,s.sold_at.strftime("%d.%m.%Y %H:%M"))
            ws.cell(ri,2,code); ws.cell(ri,3,name)
            ws.cell(ri,4,s.quantity); ws.cell(ri,5,s.sell_price)
            ws.cell(ri,6,s.sell_price*s.quantity); ws.cell(ri,7,s.profit)
        for col in ws.columns:
            mw=max((len(str(c.value or "")) for c in col),default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(mw+3,30)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        await cb.message.answer_document(
            BufferedInputFile(buf.getvalue(),f"sariosiyo_{y}_{m:02d}.xlsx"),
            caption=f"📊 {calendar.month_name[m]} {y} Excel hisobot"
        )
    except Exception as e:
        await cb.message.answer(f"❌ {e}")

# ── KAM QOLGANLAR ─────────────────────────────────────────────────────────────
@router.message(F.text == "⚠️ Kam qolganlar")
async def low_stock(msg: Message, session: AsyncSession):
    prods = await get_low_stock(session)
    adm   = await _adm(msg, session)
    if not prods:
        await msg.answer("✅ Barcha mahsulotlar yetarli!", reply_markup=main_kb(admin=adm)); return
    txt = f"⚠️ <b>Kam qolgan ({len(prods)} ta):</b>\n\n"
    for p in prods:
        icon = "🔴" if p.stock==0 else "🟡"
        txt += f"{icon} <code>{p.code}</code> — {p.name}: <b>{p.stock}</b> {p.unit}\n"
    await msg.answer(txt, parse_mode="HTML")

# ── BARCHA MAHSULOTLAR ────────────────────────────────────────────────────────
@router.message(F.text == "📋 Mahsulotlar")
async def all_prods(msg: Message, session: AsyncSession):
    if not await _adm(msg, session): return
    r = await session.execute(
        select(Product).where(Product.is_active==True).order_by(Product.category,Product.code)
    )
    prods = r.scalars().all()
    if not prods: await msg.answer("Mahsulotlar yo'q."); return
    cats = defaultdict(list)
    for p in prods: cats[p.category or "Boshqa"].append(p)
    txt = f"📋 <b>Jami: {len(prods)} ta</b>\n\n"
    for cat,items in cats.items():
        txt += f"<b>— {cat} —</b>\n"
        for p in items:
            icon = "🔴" if p.stock<=p.min_stock else "🟢"
            txt += f"{icon} <code>{p.code}</code> {p.name[:20]} — {p.stock}\n"
        txt += "\n"
        if len(txt)>3800:
            await msg.answer(txt, parse_mode="HTML"); txt=""
    if txt.strip(): await msg.answer(txt, parse_mode="HTML")

# ── REYTING ───────────────────────────────────────────────────────────────────
@router.message(F.text == "🏆 Reyting")
async def rating(msg: Message, session: AsyncSession):
    if not await _adm(msg, session): return
    now=datetime.now()
    r=await session.execute(
        select(Worker.name,func.count(Sale.id).label("c"),
               func.sum(Sale.quantity).label("q"),
               func.sum(Sale.sell_price*Sale.quantity).label("r"),
               func.sum(Sale.profit).label("p"))
        .outerjoin(Sale,Sale.worker_id==Worker.id)
        .where(Worker.is_active==True,
               extract("year",Sale.sold_at)==now.year,
               extract("month",Sale.sold_at)==now.month)
        .group_by(Worker.id).order_by(func.sum(Sale.sell_price*Sale.quantity).desc())
    )
    medals=["🥇","🥈","🥉"]
    txt=f"🏆 <b>Reyting — {calendar.month_name[now.month]}</b>\n\n"
    for i,row in enumerate(r.all()):
        m=medals[i] if i<3 else f"{i+1}."
        txt+=(f"{m} <b>{row.name}</b>\n"
              f"   📦 {row.c or 0} sotuv | {row.q or 0} dona\n"
              f"   💵 {fmt(row.r or 0)} | 📈 {fmt(row.p or 0)} so'm\n\n")
    if not r.all(): txt+="Bu oy hali sotuv yo'q."
    await msg.answer(txt, parse_mode="HTML")

# ── EKSPORT ───────────────────────────────────────────────────────────────────
@router.message(F.text == "📦 Eksport")
async def export_products(msg: Message, session: AsyncSession):
    if not await _adm(msg, session): return
    await msg.answer("⏳ Excel tayyorlanmoqda...")
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment
        from openpyxl.utils import get_column_letter
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Mahsulotlar"
        RED=PatternFill("solid",fgColor="DC1E32")
        hdrs=["Kod","Barcode","Nomi","Kategoriya","Rang","O'lcham","Birlik","Stok","Min","Tannarx","Sotish","Foyda%","Public URL"]
        for c,h in enumerate(hdrs,1):
            cell=ws.cell(1,c,h); cell.font=Font(bold=True,color="FFFFFF",size=9)
            cell.fill=RED; cell.alignment=Alignment(horizontal="center")
        r=await session.execute(
            select(Product).where(Product.is_active==True).order_by(Product.category,Product.name)
        )
        for ri,p in enumerate(r.scalars().all(),2):
            pct=((p.sell_price-p.buy_price)/p.buy_price*100) if p.buy_price>0 else 0
            vals=[p.code,p.barcode,p.name,p.category,p.color,p.size,p.unit,
                  p.stock,p.min_stock,p.buy_price,p.sell_price,round(pct,1),public_url(p)]
            for c,v in enumerate(vals,1): ws.cell(ri,c,v)
            if p.stock<=p.min_stock:
                for c in range(1,len(hdrs)+1):
                    ws.cell(ri,c).fill=PatternFill("solid",fgColor="FFCCCC")
        for col in ws.columns:
            mw=max((len(str(c.value or "")) for c in col),default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(mw+3,30)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        await msg.answer_document(
            BufferedInputFile(buf.getvalue(),f"mahsulotlar_{datetime.now():%Y%m%d}.xlsx"),
            caption=f"📦 Mahsulotlar ro'yxati — {datetime.now():%d.%m.%Y}"
        )
    except Exception as e:
        await msg.answer(f"❌ {e}")

# ── SPISANIE ──────────────────────────────────────────────────────────────────
@router.message(F.text == "🗑 Spisanie")
async def writeoff_start(msg: Message, state: FSMContext, session: AsyncSession):
    if not await _adm(msg, session): return
    await state.set_state(WriteoffSG.code)
    await msg.answer("🗑 Mahsulot kodi:", reply_markup=cancel_kb())

@router.message(WriteoffSG.code)
async def writeoff_code(msg: Message, state: FSMContext, session: AsyncSession):
    p = await get_product(session, msg.text)
    if not p: await msg.answer("⚠️ Topilmadi."); return
    await state.update_data(product_id=p.id)
    await state.set_state(WriteoffSG.qty)
    await msg.answer(f"🗑 <b>{p.name}</b>\nStok: {p.stock}\n\nNechta chiqariladi?", parse_mode="HTML")

@router.message(WriteoffSG.qty)
async def writeoff_qty(msg: Message, state: FSMContext, session: AsyncSession):
    try:
        qty=int(msg.text.strip()); assert qty>0
    except:
        await msg.answer("⚠️ Musbat son!"); return
    d=await state.get_data()
    p=await session.get(Product,d["product_id"])
    if qty>p.stock: await msg.answer(f"❌ Faqat {p.stock} bor!"); return
    await state.update_data(qty=qty)
    await state.set_state(WriteoffSG.reason)
    b=ReplyKeyboardBuilder()
    for r in ["Buzilgan","Yo'qolgan","Yaroqsiz","Namlik","Boshqa"]:
        b.button(text=r)
    b.button(text="❌ Bekor qilish"); b.adjust(3)
    await msg.answer("📝 Sabab:", reply_markup=b.as_markup(resize_keyboard=True))

@router.message(WriteoffSG.reason)
async def writeoff_reason(msg: Message, state: FSMContext, session: AsyncSession):
    reason=msg.text.strip()
    d=await state.get_data()
    p=await session.get(Product,d["product_id"])
    await state.update_data(reason=reason)
    await state.set_state(WriteoffSG.confirm)
    await msg.answer(
        f"🗑 <b>Tasdiqlash:</b>\n{p.name}\n-{d['qty']} {p.unit}\nSabab: {reason}",
        reply_markup=confirm_kb(), parse_mode="HTML"
    )

@router.callback_query(WriteoffSG.confirm, F.data == "yes")
async def writeoff_confirm(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    d=await state.get_data()
    p=await session.get(Product,d["product_id"])
    w=await get_worker(session,cb.from_user.id)
    sup=Supply(product_id=p.id,worker_id=w.id if w else None,
               quantity=-d["qty"],buy_price=p.buy_price,note=f"SPISANIE: {d['reason']}")
    p.stock-=d["qty"]; p.updated_at=datetime.utcnow()
    session.add_all([sup,p]); await session.commit()
    await state.clear()
    await cb.message.answer(
        f"✅ Spisanie: {p.name} -{d['qty']} {p.unit}\nQoldi: {p.stock}",
        reply_markup=main_kb(), parse_mode="HTML"
    )
    await cb.answer("✅")
    await check_low(cb.bot,p)

# ── XODIMLAR ──────────────────────────────────────────────────────────────────
@router.message(F.text == "👥 Xodimlar")
async def workers_list(msg: Message, session: AsyncSession):
    if not await _adm(msg, session): return
    r=await session.execute(select(Worker).order_by(Worker.name))
    ws=r.scalars().all()
    b=InlineKeyboardBuilder()
    for w in ws:
        icon="👑" if w.role=="admin" else "👤"
        st="✅" if w.is_active else "❌"
        b.button(text=f"{icon}{st} {w.name}",callback_data=f"wk:{w.id}")
    b.button(text="➕ Yangi xodim",callback_data="wk_add")
    b.adjust(1)
    await msg.answer(f"👥 <b>Xodimlar ({len(ws)} ta)</b>",
                     reply_markup=b.as_markup(),parse_mode="HTML")

@router.callback_query(F.data.startswith("wk:"))
async def worker_detail(cb: CallbackQuery, session: AsyncSession):
    w=await session.get(Worker,int(cb.data.split(":")[1]))
    if not w: await cb.answer("Topilmadi."); return
    r=await session.execute(
        select(func.count(Sale.id),func.sum(Sale.sell_price*Sale.quantity))
        .where(Sale.worker_id==w.id,func.date(Sale.sold_at)==date.today())
    )
    row=r.one()
    b=InlineKeyboardBuilder()
    b.button(text="🚫 Bloklash" if w.is_active else "✅ Faollashtirish",
             callback_data=f"wk_{'block' if w.is_active else 'unblock'}:{w.id}")
    b.button(text="⬅️ Orqaga",callback_data="wk_list")
    b.adjust(2)
    await cb.message.answer(
        f"👤 <b>{w.name}</b>\n📞 {w.phone or '—'}\n🎭 {w.role}\n"
        f"📊 Bugun: {row[0] or 0} sotuv, {fmt(row[1] or 0)} so'm",
        parse_mode="HTML",reply_markup=b.as_markup()
    )
    await cb.answer()

@router.callback_query(F.data.startswith("wk_block:"))
async def worker_block(cb: CallbackQuery, session: AsyncSession):
    w=await session.get(Worker,int(cb.data.split(":")[1]))
    if w: w.is_active=False; await session.commit()
    await cb.answer(f"🚫 {w.name} bloklandi.")

@router.callback_query(F.data.startswith("wk_unblock:"))
async def worker_unblock(cb: CallbackQuery, session: AsyncSession):
    w=await session.get(Worker,int(cb.data.split(":")[1]))
    if w: w.is_active=True; await session.commit()
    await cb.answer(f"✅ {w.name} faollashtirildi.")

@router.callback_query(F.data == "wk_list")
async def wk_list(cb: CallbackQuery, session: AsyncSession):
    r=await session.execute(select(Worker).order_by(Worker.name))
    ws=r.scalars().all()
    b=InlineKeyboardBuilder()
    for w in ws:
        icon="👑" if w.role=="admin" else "👤"; st="✅" if w.is_active else "❌"
        b.button(text=f"{icon}{st} {w.name}",callback_data=f"wk:{w.id}")
    b.button(text="➕ Yangi",callback_data="wk_add"); b.adjust(1)
    await cb.message.edit_text(f"👥 Xodimlar ({len(ws)}):",reply_markup=b.as_markup())
    await cb.answer()

@router.callback_query(F.data == "wk_add")
async def wk_add(cb: CallbackQuery, state: FSMContext):
    await state.set_state(WorkerSG.tid)
    await cb.message.answer(
        "➕ <b>Yangi xodim</b>\n\nTelegram ID kiriting:\n"
        "<i>@userinfobot ga /start yuboring → ID ko'rinadi</i>",
        reply_markup=cancel_kb(),parse_mode="HTML"
    )
    await cb.answer()

@router.message(WorkerSG.tid)
async def wk_tid(msg: Message, state: FSMContext):
    try: tid=int(msg.text.strip())
    except: await msg.answer("⚠️ Faqat raqam!"); return
    await state.update_data(tid=tid)
    await state.set_state(WorkerSG.name)
    await msg.answer("📝 Ismi:")

@router.message(WorkerSG.name)
async def wk_name(msg: Message, state: FSMContext):
    await state.update_data(name=msg.text.strip())
    await state.set_state(WorkerSG.phone)
    await msg.answer("📞 Telefon (ixtiyoriy):", reply_markup=skip_kb())

@router.message(WorkerSG.phone)
async def wk_phone(msg: Message, state: FSMContext):
    await state.update_data(phone="" if msg.text=="⏭ O'tkazish" else msg.text.strip())
    await state.set_state(WorkerSG.role)
    b=ReplyKeyboardBuilder()
    b.button(text="👤 Sotuvchi"); b.button(text="👑 Admin"); b.button(text="❌ Bekor qilish")
    b.adjust(2)
    await msg.answer("🎭 Roli:", reply_markup=b.as_markup(resize_keyboard=True))

@router.message(WorkerSG.role)
async def wk_role(msg: Message, state: FSMContext):
    role="admin" if "Admin" in msg.text else "seller"
    await state.update_data(role=role)
    await state.set_state(WorkerSG.confirm)
    d=await state.get_data()
    await msg.answer(
        f"✅ <b>Tasdiqlang:</b>\nID: <code>{d['tid']}</code>\n"
        f"Ism: {d['name']}\nTel: {d.get('phone') or '—'}\nRol: {role}",
        reply_markup=confirm_kb(),parse_mode="HTML"
    )

@router.callback_query(WorkerSG.confirm, F.data == "yes")
async def wk_confirm(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    d=await state.get_data()
    ex=await session.execute(select(Worker).where(Worker.telegram_id==d["tid"]))
    if ex.scalar_one_or_none():
        await cb.message.answer("⚠️ Bu xodim allaqachon mavjud!")
        await state.clear(); return
    w=Worker(telegram_id=d["tid"],name=d["name"],phone=d.get("phone",""),role=d["role"])
    session.add(w); await session.commit()
    await state.clear()
    await cb.message.answer(f"✅ <b>{w.name}</b> qo'shildi!",reply_markup=main_kb(),parse_mode="HTML")
    await cb.answer()
    try:
        await cb.bot.send_message(
            w.telegram_id,
            f"✅ Siz <b>Sariosiyo Bot</b>ga qo'shildingiz! Rol: <b>{w.role}</b>\n/start",
            parse_mode="HTML"
        )
    except: pass

# ── O'CHIRISH ─────────────────────────────────────────────────────────────────
@router.callback_query(F.data.startswith("del:"))
async def del_confirm(cb: CallbackQuery, session: AsyncSession):
    if not await is_admin(session,cb.from_user.id): await cb.answer("⛔️"); return
    pid=cb.data.split(":")[1]
    b=InlineKeyboardBuilder()
    b.button(text="🗑 Ha, o'chir",callback_data=f"del_yes:{pid}")
    b.button(text="❌ Yo'q",callback_data=f"sp:{pid}")
    b.adjust(2)
    await cb.message.answer("⚠️ O'chirishni tasdiqlang:",reply_markup=b.as_markup())
    await cb.answer()

@router.callback_query(F.data.startswith("del_yes:"))
async def del_do(cb: CallbackQuery, session: AsyncSession):
    p=await session.get(Product,int(cb.data.split(":")[1]))
    if p: p.is_active=False; await session.commit()
    await cb.message.answer(f"🗑 <b>{p.name}</b> arxivlandi.",
                             reply_markup=main_kb(),parse_mode="HTML")
    await cb.answer()

@router.callback_query(F.data.startswith("hist:"))
async def hist_cb(cb: CallbackQuery, session: AsyncSession):
    p=await session.get(Product,int(cb.data.split(":")[1]))
    sl=await session.execute(
        select(Sale).where(Sale.product_id==p.id).order_by(Sale.sold_at.desc()).limit(10)
    )
    tot=await session.execute(
        select(func.sum(Sale.quantity),func.sum(Sale.profit)).where(Sale.product_id==p.id)
    )
    tr=tot.one()
    txt=(f"📊 <b>{p.name}</b>\n\nJami: {tr[0] or 0} {p.unit}\nFoyda: {fmt(tr[1] or 0)} so'm\n\n<b>Oxirgi:</b>\n")
    for s in sl.scalars().all():
        txt+=f"• {s.sold_at:%d.%m %H:%M} — {s.quantity} {p.unit}\n"
    await cb.message.answer(txt,parse_mode="HTML")
    await cb.answer()

# Universal
@router.callback_query(F.data == "no")
async def any_no(cb: CallbackQuery, state: FSMContext, session: AsyncSession):
    await state.clear()
    adm=await is_admin(session,cb.from_user.id)
    await cb.message.answer("❌ Bekor.",reply_markup=main_kb(admin=adm))
    await cb.answer()

@router.callback_query(F.data.startswith("back:"))
async def back_cb(cb: CallbackQuery, session: AsyncSession):
    p=await session.get(Product,int(cb.data.split(":")[1]))
    if p: await send_product(cb.message,p)
    await cb.answer()

# ════════════════════════════════════════════════════════════════
#  10. BACKGROUND TASKS
# ════════════════════════════════════════════════════════════════
async def check_low(bot: Bot, p: Product):
    if p.stock<=p.min_stock:
        icon="🔴" if p.stock==0 else "🟡"
        txt=(f"{icon} <b>Kam qoldi!</b>\n📦 {p.name}\n"
             f"🔖 {p.code}\n📊 Qoldi: <b>{p.stock} {p.unit}</b>")
        for aid in ADMIN_IDS:
            try: await bot.send_message(aid,txt,parse_mode="HTML")
            except: pass

async def daily_report(bot: Bot):
    """Har kuni soat DAILY_REPORT_HOUR da avtomatik hisobot"""
    while True:
        now=datetime.now()
        target=now.replace(hour=DAILY_REPORT_HOUR,minute=0,second=0,microsecond=0)
        if now>=target: target=target.replace(day=target.day+1)
        await asyncio.sleep((target-now).total_seconds())
        today=date.today()
        async with AsyncSessionLocal() as s:
            r=await s.execute(
                select(func.count(Sale.id),func.sum(Sale.quantity),
                       func.sum(Sale.sell_price*Sale.quantity),func.sum(Sale.profit))
                .where(func.date(Sale.sold_at)==today)
            )
            row=r.one(); cnt,qty,rev,prf=row[0] or 0,row[1] or 0,row[2] or 0.0,row[3] or 0.0
            lows=await get_low_stock(s)
        txt=(f"📊 <b>Kunlik hisobot — {today:%d.%m.%Y}</b>\n\n"
             f"📦 Sotuvlar: <b>{cnt}</b>\n🔢 Dona: <b>{qty}</b>\n"
             f"💵 Daromad: <b>{fmt(rev)} so'm</b>\n"
             f"📈 Foyda: <b>{fmt(prf)} so'm</b>")
        if lows: txt+=f"\n\n⚠️ Kam qolgan: <b>{len(lows)} ta mahsulot</b>"
        for aid in ADMIN_IDS:
            try: await bot.send_message(aid,txt,parse_mode="HTML")
            except: pass

async def daily_backup(bot: Bot):
    """Har 24 soatda DB backup yuborish"""
    while True:
        await asyncio.sleep(86400)
        try:
            now=datetime.now()
            bpath=os.path.join(BACKUP_DIR,f"backup_{now:%Y%m%d_%H%M}.db")
            shutil.copy2(DB_PATH,bpath)
            with open(bpath,"rb") as f: data=f.read()
            for aid in ADMIN_IDS:
                try:
                    await bot.send_document(
                        aid,BufferedInputFile(data,f"backup_{now:%Y%m%d}.db"),
                        caption=f"💾 Backup — {now:%d.%m.%Y %H:%M}"
                    )
                except: pass
            # 7 kundan eski backuplarni o'chirish
            for fname in os.listdir(BACKUP_DIR):
                fpath=os.path.join(BACKUP_DIR,fname)
                if os.path.isfile(fpath):
                    age=(now-datetime.fromtimestamp(os.path.getmtime(fpath))).days
                    if age>7: os.remove(fpath)
        except Exception as e:
            logger.error(f"Backup xato: {e}")

# ════════════════════════════════════════════════════════════════
#  11. MAIN
# ════════════════════════════════════════════════════════════════
from aiogram import Bot, Dispatcher
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.fsm.storage.memory import MemoryStorage

async def on_startup(bot: Bot):
    # DB yaratish
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    # Admin xodimlarni qo'shish
    async with AsyncSessionLocal() as s:
        for aid in ADMIN_IDS:
            ex=await s.execute(select(Worker).where(Worker.telegram_id==aid))
            if not ex.scalar_one_or_none():
                s.add(Worker(telegram_id=aid,name="Admin",role="admin"))
        await s.commit()
    me=await bot.get_me()
    logger.info(f"✅ Bot: @{me.username}")
    logger.info(f"✅ Admin IDs: {ADMIN_IDS}")
    logger.info(f"✅ Kanal: {CHANNEL_ID}")
    logger.info(f"✅ Skaner: {WEBAPP_PUBLIC_URL}")
    logger.info(f"✅ Public: {WEBAPP_PUBLIC_URL}")
    logger.info(f"✅ API: {WEBAPP_PUBLIC_URL}")
    logger.info(f"✅ Admin parollar: {ADMIN_PASSWORDS}")

    # Avtomatik vazifalar
    asyncio.create_task(daily_report(bot))
    asyncio.create_task(daily_backup(bot))
    logger.info(f"✅ Kunlik hisobot: {DAILY_REPORT_HOUR}:00 da")
    logger.info(f"✅ Backup: har 24 soatda")

    # Startda adminga xabar
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(
                aid,
                f"🚀 <b>Sariosiyo Bot v7.0 ishga tushdi!</b>\n\n"
                f"🌐 Skaner: {WEBAPP_PUBLIC_URL}/scanner\n"
                f"📱 Public: {WEBAPP_PUBLIC_URL}/p/SAR-BARCODE\n"
                f"📊 API: {WEBAPP_PUBLIC_URL}/api/product\n"
                f"⏰ Hisobot: {DAILY_REPORT_HOUR}:00",
                parse_mode="HTML"
            )
        except: pass

async def main():
    if not BOT_TOKEN:
        raise SystemExit("❌ BOT_TOKEN topilmadi! .env faylini to'ldiring.")

    await start_http_server()

    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp  = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    dp.startup.register(on_startup)

    logger.info("🚀 Bot v7.0 ishga tushmoqda...")
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == "__main__":
    asyncio.run(main())